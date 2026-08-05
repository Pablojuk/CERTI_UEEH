# -*- coding: utf-8 -*-
"""
Script de procesamiento académico y generación de boletines PDF para la UEEH.
Cruza las notas de archivos Excel trimestrales y supletorios.
"""

from __future__ import annotations
from copy import copy
import os
import sys
import json
import shutil

try:
    sys.stdin.reconfigure(encoding='utf-8')
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')
except Exception:
    pass
import argparse
import math
from decimal import Decimal, InvalidOperation, ROUND_DOWN
import pandas as pd
import numpy as np
from pathlib import Path
from html import escape as escapar_html, unescape as desescapar_html
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from catalogo_asignaturas import (
    CATALOGO_ASIGNATURAS,
    clasificar_asignatura,
    convertir_nota_a_escala_cualitativa,
    grado_usa_escala_cualitativa,
    grados_equivalentes,
    metadatos_asignatura,
    normalizar_grado,
    normalizar_texto_asignatura,
    orden_asignatura,
    tipo_asignatura,
)

# ReportLab imports
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

COLUMNA_MAXIMA_ACADEMICA = column_index_from_string("X")
ENCABEZADO_EVALUACION_COMPORTAMENTAL = "EVALUACION COMPORTAMENTAL"
GRADOS_SIN_EVALUACION_COMPORTAMENTAL = {"INICIAL_1", "INICIAL_2", "EGB_1"}
CAPACIDAD_MATERIAS_FORMATO_BGU3 = 16
MAX_EXCEL_BYTES = 25 * 1024 * 1024
FILA_ENCABEZADOS_FORMATO_BGU3 = 9
PRIMERA_COLUMNA_MATERIAS_FORMATO_BGU3 = column_index_from_string("C")
VALORES_VACIOS_TEXTO = {"", "nan", "none", "null", "undefined"}
CAMPOS_METADATOS_ASIGNATURA = (
    "tipo",
    "categoria",
    "es_optativa_bgu3",
    "presentacion_certificado",
    "permite_supletorio",
    "orden",
)
DEBUG_ACTIVO = os.environ.get("CERTI_DEBUG") == "1"


def registrar_debug(mensaje):
    """Emite diagnósticos sin datos académicos únicamente cuando se habilita explícitamente."""
    if DEBUG_ACTIVO:
        print(str(mensaje), file=sys.stderr)


def leer_stdin_utf8(stdin=None):
    """Lee el contrato IPC de Python como bytes UTF-8, también dentro de PyInstaller."""
    flujo = stdin if stdin is not None else sys.stdin
    flujo_binario = getattr(flujo, "buffer", None)
    contenido = flujo_binario.read() if flujo_binario is not None else flujo.read()
    if isinstance(contenido, bytes):
        return contenido.decode("utf-8-sig")
    return str(contenido)


def escribir_json_stdout(datos, stdout=None):
    """Escribe una única respuesta JSON UTF-8 sin depender de la página de códigos."""
    flujo = stdout if stdout is not None else sys.stdout
    texto = json.dumps(datos, ensure_ascii=False) + "\n"
    flujo_binario = getattr(flujo, "buffer", None)
    if flujo_binario is not None:
        flujo_binario.write(texto.encode("utf-8"))
        flujo_binario.flush()
        return
    flujo.write(texto)
    flujo.flush()


def es_segmento_ruta_seguro(valor, maximo=180):
    texto = str(valor or "").strip()
    return bool(
        texto
        and len(texto) <= maximo
        and texto not in {".", ".."}
        and not any(caracter in texto for caracter in ("/", "\\", "\0"))
        and not any(ord(caracter) < 32 or ord(caracter) == 127 for caracter in texto)
    )


def resolver_ruta_hija(directorio_base, *segmentos):
    base = Path(directorio_base).resolve()
    destino = base.joinpath(*map(str, segmentos)).resolve()
    try:
        destino.relative_to(base)
    except ValueError:
        return None
    return destino


def validar_archivo_excel(ruta):
    try:
        archivo = Path(ruta)
        return bool(
            archivo.is_absolute()
            and archivo.suffix.lower() in {".xlsx", ".xls"}
            and archivo.is_file()
            and not archivo.is_symlink()
            and 0 < archivo.stat().st_size <= MAX_EXCEL_BYTES
        )
    except (OSError, TypeError, ValueError):
        return False


def optativas_bgu3_catalogo() -> list[dict]:
    """Devuelve las optativas autorizadas de 3.º BGU en su orden oficial."""
    return sorted(
        (
            entrada
            for entrada in CATALOGO_ASIGNATURAS
            if entrada.get("es_optativa_bgu3") is True
        ),
        key=lambda entrada: int(entrada.get("orden", 999)),
    )


def generar_formato_bgu3_con_optativas(
    origen,
    destino,
    optativas,
):
    """Genera una copia del formato de 3.º BGU con las optativas seleccionadas."""
    ruta_origen = Path(origen).resolve()
    ruta_destino = Path(destino).resolve()
    if not validar_archivo_excel(ruta_origen):
        raise ValueError("No se encontró la plantilla original de 3.º BGU.")
    if ruta_origen == ruta_destino:
        raise ValueError("La plantilla original es de solo lectura.")

    autorizadas = optativas_bgu3_catalogo()
    autorizadas_por_nombre = {
        entrada["nombre"]: entrada
        for entrada in autorizadas
    }
    seleccion = [
        str(nombre).strip()
        for nombre in (optativas if isinstance(optativas, list) else [])
        if str(nombre).strip()
    ]
    if not seleccion:
        raise ValueError("Seleccione al menos una optativa de 3.º BGU.")
    if (
        len(set(seleccion)) != len(seleccion)
        or any(nombre not in autorizadas_por_nombre for nombre in seleccion)
    ):
        raise ValueError("La selección contiene optativas no autorizadas.")
    seleccion = [
        entrada["nombre"]
        for entrada in autorizadas
        if entrada["nombre"] in seleccion
    ]

    materias_fijas_catalogo = sum(
        1
        for entrada in CATALOGO_ASIGNATURAS
        if (
            "BGU_3" in entrada.get("grados", [])
            and entrada.get("es_optativa_bgu3") is not True
        )
    ) + 1  # Evaluación comportamental no está en el catálogo.
    total_materias = materias_fijas_catalogo + len(seleccion)
    if total_materias > CAPACIDAD_MATERIAS_FORMATO_BGU3:
        raise ValueError(
            "La selección supera la capacidad de "
            f"{CAPACIDAD_MATERIAS_FORMATO_BGU3} asignaturas del formato."
        )

    ruta_destino.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(ruta_origen, ruta_destino)
    libro = load_workbook(ruta_destino)
    hoja = libro.active
    ultima_columna = (
        PRIMERA_COLUMNA_MATERIAS_FORMATO_BGU3
        + CAPACIDAD_MATERIAS_FORMATO_BGU3
        - 1
    )
    nombres_optativas_normalizados = {
        normalizar_texto_asignatura(entrada["nombre"])
        for entrada in autorizadas
    }

    columnas_originales = {}
    for columna in range(
        PRIMERA_COLUMNA_MATERIAS_FORMATO_BGU3,
        ultima_columna + 1,
    ):
        letra = get_column_letter(columna)
        dimension = hoja.column_dimensions[letra]
        columnas_originales[columna] = {
            "ancho": dimension.width,
            "oculta": dimension.hidden,
            "mejor_ajuste": dimension.bestFit,
            "nivel": dimension.outlineLevel,
            "celdas": [
                {
                    "valor": hoja.cell(fila, columna).value,
                    "estilo": copy(hoja.cell(fila, columna)._style),
                    "comentario": copy(hoja.cell(fila, columna).comment),
                    "hipervinculo": copy(hoja.cell(fila, columna).hyperlink),
                }
                for fila in range(FILA_ENCABEZADOS_FORMATO_BGU3, hoja.max_row + 1)
            ],
        }

    columnas_optativas = []
    materias_fijas = []
    posicion_primera_optativa = None
    for columna in range(
        PRIMERA_COLUMNA_MATERIAS_FORMATO_BGU3,
        ultima_columna + 1,
    ):
        encabezado = mostrar_valor(
            hoja.cell(FILA_ENCABEZADOS_FORMATO_BGU3, columna).value
        )
        if not encabezado:
            continue
        es_optativa = (
            normalizar_texto_asignatura(encabezado)
            in nombres_optativas_normalizados
        )
        if es_optativa:
            if posicion_primera_optativa is None:
                posicion_primera_optativa = len(materias_fijas)
            columnas_optativas.append(columna)
        else:
            materias_fijas.append((encabezado, columna))

    if len(materias_fijas) != materias_fijas_catalogo:
        raise ValueError(
            "La plantilla de 3.º BGU no contiene la cantidad esperada "
            "de materias generales, Cívica y Evaluación Comportamental."
        )
    if len(seleccion) > len(columnas_optativas):
        raise ValueError(
            "La plantilla no tiene suficientes columnas reservadas para "
            "las optativas seleccionadas."
        )

    punto_insercion = (
        posicion_primera_optativa
        if posicion_primera_optativa is not None
        else len(materias_fijas)
    )
    materias_generadas = [
        *materias_fijas[:punto_insercion],
        *[
            (nombre, columnas_optativas[indice])
            for indice, nombre in enumerate(seleccion)
        ],
        *materias_fijas[punto_insercion:],
    ]

    for desplazamiento, (nombre, columna_fuente) in enumerate(materias_generadas):
        columna_destino = PRIMERA_COLUMNA_MATERIAS_FORMATO_BGU3 + desplazamiento
        datos_fuente = columnas_originales[columna_fuente]
        letra_destino = get_column_letter(columna_destino)
        dimension_destino = hoja.column_dimensions[letra_destino]
        dimension_destino.width = datos_fuente["ancho"]
        dimension_destino.hidden = datos_fuente["oculta"]
        dimension_destino.bestFit = datos_fuente["mejor_ajuste"]
        dimension_destino.outlineLevel = datos_fuente["nivel"]
        for indice_fila, fila in enumerate(
            range(FILA_ENCABEZADOS_FORMATO_BGU3, hoja.max_row + 1)
        ):
            origen_celda = datos_fuente["celdas"][indice_fila]
            celda = hoja.cell(fila, columna_destino)
            celda.value = origen_celda["valor"]
            celda._style = copy(origen_celda["estilo"])
            celda.comment = copy(origen_celda["comentario"])
            celda._hyperlink = copy(origen_celda["hipervinculo"])
        hoja.cell(FILA_ENCABEZADOS_FORMATO_BGU3, columna_destino).value = nombre

    columnas_usadas = len(materias_generadas)
    for columna_destino in range(
        PRIMERA_COLUMNA_MATERIAS_FORMATO_BGU3 + columnas_usadas,
        ultima_columna + 1,
    ):
        datos_fuente = columnas_originales[
            columnas_optativas[
                (columna_destino - PRIMERA_COLUMNA_MATERIAS_FORMATO_BGU3)
                % len(columnas_optativas)
            ]
        ]
        for indice_fila, fila in enumerate(
            range(FILA_ENCABEZADOS_FORMATO_BGU3, hoja.max_row + 1)
        ):
            celda = hoja.cell(fila, columna_destino)
            celda.value = None
            celda._style = copy(datos_fuente["celdas"][indice_fila]["estilo"])
            celda.comment = None
            celda._hyperlink = None

    libro.save(ruta_destino)
    return {
        "success": True,
        "path": str(ruta_destino),
        "optativas": seleccion,
        "materiasFijas": materias_fijas_catalogo,
        "totalMaterias": total_materias,
        "capacidad": CAPACIDAD_MATERIAS_FORMATO_BGU3,
    }


def mostrar_valor(valor) -> str:
    """Representa vacíos sin convertirlos en cero ni exponer marcadores técnicos."""
    if valor is None:
        return ""
    if isinstance(valor, (float, np.floating)) and math.isnan(float(valor)):
        return ""
    texto = str(valor).strip()
    if texto.lower() in VALORES_VACIOS_TEXTO:
        return ""
    return texto


def completar_metadatos_asignatura(nombre: str, datos: dict | None = None) -> dict:
    """Combina la fuente de verdad del catálogo con metadatos ya propagados."""
    metadatos = metadatos_asignatura(nombre)
    if isinstance(datos, dict):
        for campo in CAMPOS_METADATOS_ASIGNATURA:
            if campo in datos and datos[campo] is not None:
                metadatos[campo] = datos[campo]
    return metadatos


def convertir_nota_optativa_a_escala_cualitativa(valor) -> str:
    """Compatibilidad para optativas: reutiliza la escala central y descarta entradas inválidas."""
    if not mostrar_valor(valor):
        return ""
    try:
        nota = Decimal(str(valor).replace(",", "."))
    except (InvalidOperation, ValueError, TypeError):
        return ""
    if not nota.is_finite() or nota < Decimal("1.00") or nota > Decimal("10.00"):
        return ""
    return convertir_nota_a_escala_cualitativa(nota)


def formatear_nota_numerica_certificado(valor) -> str:
    """Mantiene la presentación numérica histórica con dos decimales."""
    texto = mostrar_valor(valor)
    if not texto:
        return ""
    try:
        numero = Decimal(texto.replace(",", "."))
    except (InvalidOperation, ValueError, TypeError):
        return texto
    if not numero.is_finite():
        return ""
    return f"{numero:.2f}"


def presentar_nota_certificado(valor, grado, metadatos=None) -> str:
    """Presenta una nota sin alterar el valor usado por los cálculos académicos."""
    texto = mostrar_valor(valor)
    if not texto:
        return ""
    datos = metadatos if isinstance(metadatos, dict) else {}
    if datos.get("tipo") == "cualitativa":
        return texto
    if datos.get("presentacion_certificado") == "escala_cualitativa":
        return convertir_nota_optativa_a_escala_cualitativa(valor)
    if grado_usa_escala_cualitativa(grado):
        try:
            numero = Decimal(texto.replace(",", "."))
        except (InvalidOperation, ValueError, TypeError):
            return texto
        if numero.is_finite() and Decimal("1") <= numero <= Decimal("10"):
            return convertir_nota_a_escala_cualitativa(numero)
    return formatear_nota_numerica_certificado(valor)


def es_encabezado_evaluacion_comportamental(valor) -> bool:
    return normalizar_texto_asignatura(valor) == ENCABEZADO_EVALUACION_COMPORTAMENTAL


def curso_admite_evaluacion_comportamental(grado) -> bool:
    codigo = normalizar_grado(grado)
    return codigo not in GRADOS_SIN_EVALUACION_COMPORTAMENTAL


# 1. Funciones de cálculo académico
def truncar_2_decimales(valor: float) -> float:
    try:
        # Decimal evita que 9.20 se convierta accidentalmente en 9.19 por representación binaria.
        return float(Decimal(str(valor)).quantize(Decimal("0.01"), rounding=ROUND_DOWN))
    except (InvalidOperation, ValueError, TypeError):
        return 0.0

def calcular_promedio_anual(t1: float, t2: float, t3: float) -> float:
    return truncar_2_decimales((t1 + t2 + t3) / 3.0)

def calcular_resultado_con_supletorio(nota_final: float, nota_supletorio: float | None) -> float:
    if nota_final >= 7.0:
        return truncar_2_decimales(nota_final)
    if nota_supletorio is not None and nota_supletorio >= 7.0:
        return 7.0
    return truncar_2_decimales(nota_final)

def determinar_estado_materia(nota_final: float, nota_supletorio: float | None) -> str:
    if nota_final >= 7.0:
        return "APROBADO"
    if nota_supletorio is not None:
        # Ya tomó supletorio y no llegó a 7
        return "REPROBADO"
    return "SUPLETORIO"

# 2. Análisis e Importación de Excels
def normalizar_columna(col_name: str) -> str:
    return normalizar_texto_asignatura(col_name).lower()


class ErrorGradoExcel(ValueError):
    """El libro pertenece inequívocamente a otro curso."""


def crear_diagnostico_asignaturas() -> dict:
    return {
        "asignaturasReconocidas": [],
        "asignaturasIgnoradasPorCurso": [],
        "asignaturasNoReconocidas": [],
        "optativasDetectadas": [],
    }


def _agregar_diagnostico(diagnostico: dict, clasificacion: dict) -> None:
    estado = clasificacion.get("estado")
    if estado == "reconocida":
        reconocida = {
            "original": clasificacion["original"],
            "canonica": clasificacion["canonica"],
            "metodo": clasificacion["metodo"],
            "tipo": clasificacion.get("tipo", "cuantitativa"),
            "categoria": clasificacion.get("categoria"),
            "es_optativa_bgu3": clasificacion.get("es_optativa_bgu3", False),
            "presentacion_certificado": clasificacion.get("presentacion_certificado"),
            "permite_supletorio": clasificacion.get("permite_supletorio", True),
            "orden": clasificacion.get("orden", 999),
        }
        diagnostico["asignaturasReconocidas"].append(reconocida)
        if (
            clasificacion.get("es_optativa_bgu3")
            and not any(
                item["canonica"] == clasificacion["canonica"]
                for item in diagnostico["optativasDetectadas"]
            )
        ):
            diagnostico["optativasDetectadas"].append({
                "original": clasificacion["original"],
                "canonica": clasificacion["canonica"],
                "metodo": clasificacion["metodo"],
                "presentacion": clasificacion.get("presentacion_certificado"),
                "permite_supletorio": clasificacion.get("permite_supletorio", True),
            })
    elif estado == "ignorada_por_curso":
        diagnostico["asignaturasIgnoradasPorCurso"].append({
            "original": clasificacion["original"],
            "canonica": clasificacion["canonica"],
            "motivo": clasificacion["motivo"],
        })
    else:
        diagnostico["asignaturasNoReconocidas"].append({
            "original": clasificacion["original"],
            "motivo": "No coincide de forma confiable con el catálogo oficial",
        })

def detectar_columnas(df: pd.DataFrame) -> tuple[str | None, str | None, list[str]]:
    id_col = None
    name_col = None
    subject_cols = []
    
    ignore_keywords = [
        "nro", "n.", "no.", "lista", "numero", "promedio", "total", 
        "estado", "observacion", "comentario", "firma", "obs"
    ]
    
    for col in df.columns:
        norm = normalizar_columna(col)
        # Identificar columna ID
        if any(k in norm for k in ["cedula", "identificacion", "id", "codigo"]):
            if id_col is None:
                id_col = col
            continue
        # Identificar columna Nombre
        if any(k in norm for k in ["estudiante", "nombre", "apellidos", "alumno", "listado"]):
            if name_col is None:
                name_col = col
            continue
            
    # Asignaturas: Columnas numéricas que no se ignoran
    for col in df.columns:
        if col == id_col or col == name_col:
            continue
        norm = normalizar_columna(col)
        if any(k in norm for k in ignore_keywords):
            continue
        
        # Verificar si la columna tiene valores numéricos
        numeric_series = pd.to_numeric(df[col], errors='coerce')
        if numeric_series.notna().any():
            subject_cols.append(col)
            
    return id_col, name_col, subject_cols

def obtener_hoja_principal(wb):
    """Obtiene la hoja oficial principal del formato de notas."""
    if "Reporte Periodo" in wb.sheetnames:
        return wb["Reporte Periodo"]
    return None


def _celda_con_valor_combinado(sheet, cell_ref):
    """Devuelve la celda directa o la celda superior izquierda del rango combinado."""
    cell = sheet[cell_ref]
    if cell.value not in (None, ""):
        return cell

    for merged_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
    return cell


def _valor_a_texto(value, cell=None):
    """Convierte valores de Excel a texto limpio, respetando formatos con ceros."""
    if value is None:
        return ""

    if isinstance(value, float) and value.is_integer():
        value = int(value)

    if isinstance(value, int):
        number_format = getattr(cell, "number_format", "") if cell is not None else ""
        if number_format and set(number_format) == {"0"} and len(number_format) > len(str(value)):
            return str(value).zfill(len(number_format)).strip()
        return str(value).strip()

    return str(value).strip()


def obtener_valor_celda(sheet, cell_ref, fallback_refs=None):
    """
    Lee una celda, revisa rangos combinados si está vacía y luego prueba fallbacks.
    Devuelve siempre texto limpio.
    """
    refs = [cell_ref, *(fallback_refs or [])]
    for ref in refs:
        cell = _celda_con_valor_combinado(sheet, ref)
        value = _valor_a_texto(cell.value, cell)
        if value:
            return value
    return ""

def cargar_excel_datos(file_path: str, grado_esperado: str | None = None, diagnostico: dict | None = None) -> dict[str, dict]:
    """
    Lee el excel buscando la hoja oficial 'Reporte Periodo'.
    Busca a partir de la fila donde dice 'LISTADO' en la columna A
    y debajo de 'CEDULA' en la columna B para extraer el nombre y la cédula/ID.
    Retorna un diccionario indexado por cédula/nombre de los estudiantes,
    con 'notas' vacío por ahora.
    """
    diagnostico = diagnostico if diagnostico is not None else crear_diagnostico_asignaturas()
    if not validar_archivo_excel(file_path):
        registrar_debug("[cargar_excel_datos] Ruta inválida o inexistente.")
        return {}
        
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = obtener_hoja_principal(wb)
        if sheet is None:
            registrar_debug("[cargar_excel_datos] No existe la hoja obligatoria.")
            return {}
        registrar_debug("[cargar_excel_datos] Hoja principal localizada.")

        grado_excel = obtener_valor_celda(sheet, "B3")
        if grado_esperado:
            if not grado_excel:
                raise ErrorGradoExcel("El archivo no indica el grado en la celda B3.")
            if not grados_equivalentes(grado_excel, grado_esperado):
                raise ErrorGradoExcel(
                    f"El archivo corresponde a {grado_excel}, pero el curso seleccionado es {grado_esperado}."
                )
        
        # Buscar la fila de encabezado
        start_row = None
        for r in range(1, 100):
            val_a = sheet.cell(row=r, column=1).value
            val_b = sheet.cell(row=r, column=2).value
            if val_a and val_b:
                str_a = str(val_a).strip().upper()
                str_b = str(val_b).strip().upper()
                if 'LISTADO' in str_a and 'CEDULA' in str_b:
                    start_row = r
                    break
                    
        if start_row is None:
            # Fallback
            for r in range(1, 100):
                val_a = sheet.cell(row=r, column=1).value
                if val_a and 'LISTADO' in str(val_a).strip().upper():
                    start_row = r
                    break
                    
        if start_row is None:
            return {}
            
        ignore_subject_keywords = {"promedio", "total", "observacion", "observación", "estado", "firma", "comportamiento"}
        
        # Detectar si es formato de materia única (donde Col C es T1, Col E es T2, Col G es T3)
        es_materia_unica = False
        header_c = str(sheet.cell(row=start_row, column=3).value or "").strip()
        
        if sheet.max_column >= 5:
            header_d = str(sheet.cell(row=start_row, column=4).value or "").strip().upper()
            header_e = str(sheet.cell(row=start_row, column=5).value or "").strip().upper()
            if any(k in header_d for k in ["EXAM", "PROY", "EVAL"]) or any(k in header_e for k in ["TRIMESTRE 2", "T2", "EXAM", "EVAL"]):
                es_materia_unica = True
                
        subject_columns = []
        canonicas_agregadas = set()
        if es_materia_unica:
            periodo_detectado = str(sheet.cell(row=6, column=2).value or "").strip().upper()
            col_to_read = 3
            if "TRIMESTRE 2" in periodo_detectado or "T2" in periodo_detectado or "SEGUNDO" in periodo_detectado:
                col_to_read = 5
            elif "TRIMESTRE 3" in periodo_detectado or "T3" in periodo_detectado or "TERCER" in periodo_detectado:
                col_to_read = 7
            
            subject_name = header_c if header_c else "Materia"
            clasificacion = clasificar_asignatura(subject_name, grado_esperado)
            _agregar_diagnostico(diagnostico, clasificacion)
            if clasificacion["estado"] == "reconocida":
                subject_columns.append({
                    "columna": col_to_read,
                    "nombre": clasificacion["canonica"],
                    "tipo": clasificacion.get("tipo", "cuantitativa"),
                    "metadatos": {
                        campo: clasificacion.get(campo)
                        for campo in CAMPOS_METADATOS_ASIGNATURA
                    },
                })
                canonicas_agregadas.add(clasificacion["canonica"])
            registrar_debug("[cargar_excel_datos] Formato de materia única detectado.")
        else:
            # Formato estándar multiasignatura. La zona académica oficial termina en X.
            comportamiento_col = None
            columna_final = min(sheet.max_column, COLUMNA_MAXIMA_ACADEMICA)
            for col in range(3, columna_final + 1):
                header_value = sheet.cell(row=start_row, column=col).value
                if header_value is None or str(header_value).strip() == "":
                    continue
                subject_name = str(header_value).strip()
                if es_encabezado_evaluacion_comportamental(subject_name):
                    comportamiento_col = col
                    continue
                subject_norm = normalizar_columna(subject_name)
                if any(keyword in subject_norm for keyword in ignore_subject_keywords):
                    continue
                clasificacion = clasificar_asignatura(subject_name, grado_esperado)
                _agregar_diagnostico(diagnostico, clasificacion)
                if clasificacion["estado"] != "reconocida":
                    continue
                canonica = clasificacion["canonica"]
                if canonica in canonicas_agregadas:
                    continue
                subject_columns.append({
                    "columna": col,
                    "nombre": canonica,
                    "tipo": clasificacion.get("tipo", "cuantitativa"),
                    "metadatos": {
                        campo: clasificacion.get(campo)
                        for campo in CAMPOS_METADATOS_ASIGNATURA
                    },
                })
                canonicas_agregadas.add(canonica)

        registrar_debug(
            f"[cargar_excel_datos] Cantidad de asignaturas detectadas: {len(subject_columns)}"
        )

        records = {}
        # Leer a partir de la fila siguiente
        for r in range(start_row + 1, sheet.max_row + 1):
            val_a = sheet.cell(row=r, column=1).value
            val_b = sheet.cell(row=r, column=2).value
            
            # Detenerse en fila vacía o sin nombre en columna A
            if val_a is None or str(val_a).strip() == "":
                break
                
            nombre = str(val_a).strip()
            cedula = ""
            if val_b is not None:
                val_b_str = _valor_a_texto(val_b, sheet.cell(row=r, column=2))
                if val_b_str.endswith('.0'):
                    val_b_str = val_b_str[:-2]
                if val_b_str.isdigit() and len(val_b_str) == 9:
                    val_b_str = "0" + val_b_str
                cedula = val_b_str
                
            if not cedula:
                cedula = nombre

            notas = {}
            tipos_asignaturas = {}
            metadatos_asignaturas = {}
            for subject in subject_columns:
                col = subject["columna"]
                subject_name = subject["nombre"]
                subject_type = subject["tipo"]
                tipos_asignaturas[subject_name] = subject_type
                metadatos_asignaturas[subject_name] = completar_metadatos_asignatura(
                    subject_name,
                    subject.get("metadatos"),
                )
                grade_value = sheet.cell(row=r, column=col).value
                if not mostrar_valor(grade_value):
                    notas[subject_name] = None
                    continue
                if subject_type == "cualitativa":
                    notas[subject_name] = str(grade_value).strip()
                    continue
                try:
                    grade_number = float(grade_value)
                    notas[subject_name] = None if math.isnan(grade_number) else truncar_2_decimales(grade_number)
                except (TypeError, ValueError):
                    notas[subject_name] = None

            evaluacion_comportamental = ""
            if not es_materia_unica and comportamiento_col is not None:
                evaluacion_comportamental = mostrar_valor(sheet.cell(row=r, column=comportamiento_col).value)
                
            records[cedula] = {
                "nombre": nombre,
                "cedula": cedula,
                "notas": notas,
                "tipos_asignaturas": tipos_asignaturas,
                "metadatos_asignaturas": metadatos_asignaturas,
                "evaluacion_comportamental": evaluacion_comportamental,
            }
            
        return records
    except ErrorGradoExcel:
        raise
    except Exception:
        registrar_debug("[cargar_excel_datos] El archivo no pudo procesarse.")
        return {}

def consolidar_estudiantes(t1_path, t2_path, t3_path, su_path, grado_esperado=None) -> list[dict]:
    t1_data = cargar_excel_datos(t1_path, grado_esperado)
    t2_data = cargar_excel_datos(t2_path, grado_esperado)
    t3_data = cargar_excel_datos(t3_path, grado_esperado)
    su_data = cargar_excel_datos(su_path, grado_esperado)
    
    # Unir todas las cédulas/estudiantes únicos
    all_keys = set(t1_data.keys()) | set(t2_data.keys()) | set(t3_data.keys()) | set(su_data.keys())
    
    lista_consolidada = []
    
    for key in all_keys:
        # Recuperar nombres
        nombre = ""
        for src in [t1_data, t2_data, t3_data, su_data]:
            if key in src:
                nombre = src[key]["nombre"]
                break
        
        # Consolidar asignaturas
        all_subjects = set()
        for src in [t1_data, t2_data, t3_data, su_data]:
            if key in src:
                all_subjects.update(src[key]["notas"].keys())
                
        subjects_grades = {}
        for sub in sorted(all_subjects, key=orden_asignatura):
            g1 = t1_data.get(key, {}).get("notas", {}).get(sub)
            g2 = t2_data.get(key, {}).get("notas", {}).get(sub)
            g3 = t3_data.get(key, {}).get("notas", {}).get(sub)
            metadatos = next(
                (
                    src[key].get("metadatos_asignaturas", {}).get(sub)
                    for src in [t1_data, t2_data, t3_data, su_data]
                    if key in src and src[key].get("metadatos_asignaturas", {}).get(sub)
                ),
                None,
            )
            metadatos = completar_metadatos_asignatura(sub, metadatos)
            tipo = metadatos["tipo"]

            if tipo == "cualitativa":
                subjects_grades[sub] = {
                    **metadatos,
                    "t1": mostrar_valor(g1),
                    "t2": mostrar_valor(g2),
                    "t3": mostrar_valor(g3),
                    "promedio_anual": None,
                    "supletorio": None,
                    "nota_final": None,
                    "estado": None,
                }
                continue

            permite_supletorio = metadatos["permite_supletorio"]
            su_grade = (
                su_data.get(key, {}).get("notas", {}).get(sub, None)
                if su_path and permite_supletorio
                else None
            )
            if any(nota is None for nota in (g1, g2, g3)):
                p_anual = None
                nota_final = None
                estado = "PENDIENTE DE NOTAS EXTERNAS" if permite_supletorio else None
            else:
                p_anual = calcular_promedio_anual(g1, g2, g3)
                if permite_supletorio:
                    nota_final = calcular_resultado_con_supletorio(p_anual, su_grade)
                    estado = determinar_estado_materia(nota_final, su_grade)
                else:
                    nota_final = p_anual
                    estado = None
            
            subjects_grades[sub] = {
                **metadatos,
                "t1": g1,
                "t2": g2,
                "t3": g3,
                "promedio_anual": p_anual,
                "supletorio": su_grade,
                "nota_final": nota_final,
                "estado": estado
            }
            
        # Calcular promedio general
        materias_cuantitativas = [
            val for val in subjects_grades.values()
            if val.get("tipo", "cuantitativa") != "cualitativa"
            and val.get("permite_supletorio", True)
        ]
        tiene_pendiente = any(
            val["estado"] == "PENDIENTE DE NOTAS EXTERNAS"
            for val in materias_cuantitativas
        )
        if materias_cuantitativas and not tiene_pendiente:
            final_grades_list = [
                val["nota_final"] for val in materias_cuantitativas
                if val["nota_final"] is not None
            ]
            prom_general = sum(final_grades_list) / len(final_grades_list)
        else:
            prom_general = None
            
        # Determinar estado general del estudiante
        tiene_supletorio = any(val["estado"] == "SUPLETORIO" for val in materias_cuantitativas)
        tiene_reprobado = any(val["estado"] == "REPROBADO" for val in materias_cuantitativas)
        
        if not materias_cuantitativas or tiene_pendiente:
            estado_general = "PENDIENTE DE NOTAS EXTERNAS"
        elif tiene_reprobado:
            estado_general = "REPROBADO"
        elif tiene_supletorio:
            estado_general = "SUPLETORIO"
        else:
            estado_general = "APROBADO"
            
        lista_consolidada.append({
            "cedula": key if len(key) < 15 else "S/C",  # Mostrar sin cédula si se usó el nombre largo como ID
            "id_real": key,
            "nombre": nombre,
            "promedio": prom_general,
            "estado": estado_general,
            "materias": subjects_grades,
            "evaluacion_comportamental": {
                "T1": mostrar_valor(t1_data.get(key, {}).get("evaluacion_comportamental")),
                "T2": mostrar_valor(t2_data.get(key, {}).get("evaluacion_comportamental")),
                "T3": mostrar_valor(t3_data.get(key, {}).get("evaluacion_comportamental")),
            },
        })
        
    # Ordenar por nombre
    lista_consolidada.sort(key=lambda x: x["nombre"])
    return lista_consolidada

# 3. Generación del Reporte PDF con ReportLab
def generar_boletin_pdf(datos_consolidados, institucion_info, logos_paths, output_path):
    # Configuración de hoja y márgenes
    doc = SimpleDocTemplate(
        output_path,
        pagesize=letter,
        leftMargin=36,
        rightMargin=36,
        topMargin=36,
        bottomMargin=36
    )
    
    styles = getSampleStyleSheet()
    
    # Colores personalizados (Estilo Premium)
    c_primary = colors.HexColor("#312e81")   # Indigo oscuro
    c_text = colors.HexColor("#1e293b")      # Slate 800
    c_light = colors.HexColor("#f8fafc")     # Slate 50
    c_border = colors.HexColor("#cbd5e1")    # Slate 300
    
    # Estilos de Texto
    styles.add(ParagraphStyle(
        name='MainTitle',
        fontName='Helvetica-Bold',
        fontSize=12,
        leading=14,
        textColor=c_primary,
        alignment=1 # Center
    ))
    
    styles.add(ParagraphStyle(
        name='SubTitle',
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=11,
        textColor=colors.HexColor("#475569"),
        alignment=1 # Center
    ))
    
    styles.add(ParagraphStyle(
        name='TableText',
        fontName='Helvetica',
        fontSize=8,
        leading=10,
        textColor=c_text
    ))

    styles.add(ParagraphStyle(
        name='TableTextBold',
        fontName='Helvetica-Bold',
        fontSize=8,
        leading=10,
        textColor=c_text
    ))

    styles.add(ParagraphStyle(
        name='TableHeader',
        fontName='Helvetica-Bold',
        fontSize=8,
        leading=10,
        textColor=colors.white
    ))

    story = []
    
    for i, est in enumerate(datos_consolidados):
        # 1. Encabezado con Logos
        logo_left = None
        logo_right = None
        
        # Validar y cargar imagen izquierda
        if logos_paths.get("logo1") and os.path.exists(logos_paths["logo1"]):
            try:
                logo_left = Image(logos_paths["logo1"], width=55, height=55)
            except:
                pass
        # Validar y cargar imagen derecha
        if logos_paths.get("logo2") and os.path.exists(logos_paths["logo2"]):
            try:
                logo_right = Image(logos_paths["logo2"], width=55, height=55)
            except:
                pass
                
        # Texto central del encabezado
        text_encabezado = [
            Paragraph(escapar_html(mostrar_valor(institucion_info["nombre"]).upper()), styles['MainTitle']),
            Spacer(1, 2),
            Paragraph("REPORTE DE CALIFICACIONES TRIMESTRALES", styles['SubTitle']),
            Paragraph(f"Año Lectivo: {escapar_html(mostrar_valor(institucion_info['anio']))}", styles['SubTitle']),
            Paragraph(f"Código AMIE: {escapar_html(mostrar_valor(institucion_info['amie']))}", styles['SubTitle'])
        ]
        
        # Tabla del encabezado
        header_data = [[logo_left or "", text_encabezado, logo_right or ""]]
        header_table = Table(header_data, colWidths=[65, 410, 65])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        
        story.append(header_table)
        story.append(Spacer(1, 8))
        
        # 2. Información del Estudiante
        tutor_nombre = institucion_info.get("tutor") or "N/A"
        info_data = [
            [
                Paragraph(f"<b>Estudiante:</b> {escapar_html(mostrar_valor(est['nombre']))}", styles['TableText']),
                Paragraph(f"<b>Cédula:</b> {escapar_html(mostrar_valor(est['cedula']))}", styles['TableText'])
            ],
            [
                Paragraph(f"<b>Curso:</b> {escapar_html(mostrar_valor(institucion_info['grado']))}", styles['TableText']),
                Paragraph(f"<b>Paralelo:</b> {escapar_html(mostrar_valor(institucion_info['paralelo']))}", styles['TableText'])
            ],
            [
                Paragraph(f"<b>Jornada:</b> {escapar_html(mostrar_valor(institucion_info['jornada']))}", styles['TableText']),
                Paragraph(f"<b>Tutor/a:</b> {escapar_html(mostrar_valor(tutor_nombre))}", styles['TableText'])
            ]
        ]
        
        info_table = Table(info_data, colWidths=[270, 270])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), c_light),
            ('BOX', (0,0), (-1,-1), 1, c_border),
            ('INNERGRID', (0,0), (-1,-1), 0.5, c_border),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ('LEFTPADDING', (0,0), (-1,-1), 6),
        ]))
        
        story.append(info_table)
        story.append(Spacer(1, 10))
        
        # 3. Tabla de Calificaciones
        table_headers = [
            Paragraph("Asignatura / Área", styles['TableHeader']),
            Paragraph("Trim 1", styles['TableHeader']),
            Paragraph("Trim 2", styles['TableHeader']),
            Paragraph("Trim 3", styles['TableHeader']),
            Paragraph("Prom Anual", styles['TableHeader']),
            Paragraph("Supletorio", styles['TableHeader']),
            Paragraph("Nota Final", styles['TableHeader']),
            Paragraph("Estado", styles['TableHeader'])
        ]
        
        grades_data = [table_headers]
        
        def nota_pdf(valor):
            if not mostrar_valor(valor) or str(valor).strip() == "PENDIENTE":
                return None
            try:
                numero = float(valor)
                return None if math.isnan(numero) else numero
            except (TypeError, ValueError):
                return None

        for sub, grades in est["materias"].items():
            es_cualitativa = grades.get("tipo", tipo_asignatura(sub)) == "cualitativa"
            if es_cualitativa:
                t1_str = escapar_html(mostrar_valor(grades.get("t1")))
                t2_str = escapar_html(mostrar_valor(grades.get("t2")))
                t3_str = escapar_html(mostrar_valor(grades.get("t3")))
                pa_str = su_str = nf_str = ""
            else:
                t1_val = nota_pdf(grades.get('t1'))
                t2_val = nota_pdf(grades.get('t2'))
                t3_val = nota_pdf(grades.get('t3'))
                pa_val = nota_pdf(grades.get('promedio_anual'))
                su_val = nota_pdf(grades.get('supletorio'))
                nf_val = nota_pdf(grades.get('nota_final'))

                t1_str = f"{t1_val:.2f}" if t1_val is not None else ""
                t2_str = f"{t2_val:.2f}" if t2_val is not None else ""
                t3_str = f"{t3_val:.2f}" if t3_val is not None else ""
                pa_str = f"{pa_val:.2f}" if pa_val is not None else ""
                su_str = f"{su_val:.2f}" if su_val is not None else ""
                nf_str = f"{nf_val:.2f}" if nf_val is not None else ""
            
            # Estilos de color para el estado
            est_materia = "" if es_cualitativa else grades.get('estado', 'APROBADO')
            if est_materia == "APROBADO":
                est_color = colors.HexColor("#065f46") # Emerald oscuro
            elif est_materia == "SUPLETORIO":
                est_color = colors.HexColor("#92400e") # Amber oscuro
            else:
                est_color = colors.HexColor("#991b1b") # Rose oscuro
                
            grades_data.append([
                Paragraph(escapar_html(mostrar_valor(sub)), styles['TableTextBold']),
                Paragraph(t1_str, styles['TableText']),
                Paragraph(t2_str, styles['TableText']),
                Paragraph(t3_str, styles['TableText']),
                Paragraph(pa_str, styles['TableText']),
                Paragraph(su_str, styles['TableText']),
                Paragraph(nf_str, styles['TableTextBold']),
                Paragraph(
                    "" if es_cualitativa else f"<font color='{est_color}'><b>{escapar_html(mostrar_valor(est_materia))}</b></font>",
                    styles['TableText'],
                )
            ])
            
        grades_table = Table(grades_data, colWidths=[180, 50, 50, 50, 60, 55, 55, 60])
        
        # Aplicar estilos a la tabla
        t_style = [
            ('BACKGROUND', (0,0), (-1,0), c_primary),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('ALIGN', (0,1), (0,-1), 'LEFT'), # Alinear nombres de asignaturas a la izquierda
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.5, c_border),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ]
        
        # Alternar colores de filas
        for r_idx in range(1, len(grades_data)):
            if r_idx % 2 == 0:
                t_style.append(('BACKGROUND', (0, r_idx), (-1, r_idx), c_light))
                
        grades_table.setStyle(TableStyle(t_style))
        story.append(grades_table)
        story.append(Spacer(1, 10))

        if curso_admite_evaluacion_comportamental(institucion_info.get("grado")):
            evaluacion = est.get("evaluacion_comportamental") or {}
            comportamiento_data = [[
                Paragraph("<b>EVALUACIÓN COMPORTAMENTAL</b>", styles["TableTextBold"]),
                Paragraph(escapar_html(mostrar_valor(evaluacion.get("T1"))), styles["TableText"]),
                Paragraph(escapar_html(mostrar_valor(evaluacion.get("T2"))), styles["TableText"]),
                Paragraph(escapar_html(mostrar_valor(evaluacion.get("T3"))), styles["TableText"]),
            ]]
            comportamiento_table = Table(comportamiento_data, colWidths=[135, 135, 135, 135])
            comportamiento_table.setStyle(TableStyle([
                ("GRID", (0, 0), (-1, -1), 0.5, c_border),
                ("BACKGROUND", (0, 0), (0, 0), c_light),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]))
            story.append(comportamiento_table)
            story.append(Spacer(1, 10))

        asistencia = est.get("asistencia") if isinstance(est.get("asistencia"), dict) else {}
        asistencia_configurada = bool(asistencia.get("configurada"))
        def valor_asistencia(campo):
            return mostrar_valor(asistencia.get(campo)) if asistencia_configurada else ""

        asistencia_data = [
            [
                Paragraph("REGISTRO", styles["TableHeader"]),
                Paragraph("JUSTIFICACI\u00d3N", styles["TableHeader"]),
                Paragraph("INJUSTIFICADO", styles["TableHeader"]),
                Paragraph("TOTAL ASISTENCIA", styles["TableHeader"]),
            ],
            [
                Paragraph(valor_asistencia("totalFaltas"), styles["TableText"]),
                Paragraph(valor_asistencia("justificadas"), styles["TableText"]),
                Paragraph(valor_asistencia("injustificadas"), styles["TableText"]),
                Paragraph(valor_asistencia("totalAsistencia"), styles["TableTextBold"]),
            ],
        ]
        asistencia_table = Table(asistencia_data, colWidths=[135, 135, 135, 135])
        asistencia_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), c_primary),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, c_border),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(Paragraph("<b>ASISTENCIA ANUAL</b>", styles["TableTextBold"]))
        story.append(Spacer(1, 3))
        story.append(asistencia_table)
        story.append(Spacer(1, 10))
        
        # 4. Resumen Final
        status_color = "#047857" if est['estado'] == "APROBADO" else ("#b45309" if est['estado'] in ["SUPLETORIO", "PENDIENTE DE NOTAS EXTERNAS"] else "#b91c1c")
        try:
            prom_raw = est.get('promedio')
            prom_val = float(prom_raw) if prom_raw is not None else None
        except:
            prom_val = None
        prom_txt = f"{prom_val:.2f}" if prom_val is not None else ""
            
        summary_data = [
            [
                Paragraph(f"<b>PROMEDIO GENERAL DEL ESTUDIANTE:</b> {prom_txt}", styles['TableTextBold']),
                Paragraph(f"<b>ESTADO FINAL:</b> <font color='{status_color}'><b>{escapar_html(mostrar_valor(est['estado']))}</b></font>", styles['TableTextBold'])
            ]
        ]
        summary_table = Table(summary_data, colWidths=[270, 270])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#f1f5f9")),
            ('BOX', (0,0), (-1,-1), 1, c_border),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 35))
        
        # 5. Firmas (Tutor y Rector)
        firmas_data = [
            [
                Paragraph("________________________________________", styles['SubTitle']),
                Paragraph("________________________________________", styles['SubTitle'])
            ],
            [
                Paragraph(f"<b>{escapar_html(mostrar_valor(tutor_nombre))}</b><br/>Tutor / Docente", styles['SubTitle']),
                Paragraph(f"<b>{escapar_html(mostrar_valor(institucion_info['rector']))}</b><br/>Rector / Director", styles['SubTitle'])
            ]
        ]
        firmas_table = Table(firmas_data, colWidths=[270, 270])
        firmas_table.setStyle(TableStyle([
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 2),
        ]))
        
        story.append(KeepTogether([firmas_table]))
        
        # Añadir salto de página si no es el último estudiante
        if i < len(datos_consolidados) - 1:
            story.append(PageBreak())
            
    doc.build(story)


def extraer_datos_institucionales(file_path: str) -> dict:
    if not file_path or str(file_path).strip() == "":
        raise ValueError("Ruta inválida: no se recibió la ruta del archivo Excel.")

    file_path = str(file_path).strip()
    registrar_debug("[extraer_datos_institucionales] Archivo recibido.")

    if not validar_archivo_excel(file_path):
        raise ValueError("Ruta inválida: el archivo Excel no es válido o supera 25 MB.")

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        raise ValueError(f"Excel corrupto o ilegible: no se pudo abrir el archivo. Detalle: {e}") from e

    sheet = obtener_hoja_principal(wb)
    if sheet is None:
        hojas = ", ".join(wb.sheetnames) or "sin hojas"
        raise ValueError(f"Formato incorrecto: no existe la hoja obligatoria 'Reporte Periodo'. Hojas encontradas: {hojas}")

    registrar_debug("[extraer_datos_institucionales] Hoja principal localizada.")

    valores_log = {
        "A1": obtener_valor_celda(sheet, "A1"),
        "D1/D2": obtener_valor_celda(sheet, "D2", ["D1"]),
        "B7": obtener_valor_celda(sheet, "B7"),
        "B3": obtener_valor_celda(sheet, "B3"),
        "B5": obtener_valor_celda(sheet, "B5"),
        "B4": obtener_valor_celda(sheet, "B4"),
        "D3": obtener_valor_celda(sheet, "D3"),
        "D4": obtener_valor_celda(sheet, "D4"),
        "B6": obtener_valor_celda(sheet, "B6"),
    }
    datos = {
        "nombreInstitucion": valores_log["A1"],
        "nivel": "",
        "anioLectivo": valores_log["D1/D2"],
        "codigoAmie": valores_log["B7"],
        "gradoCurso": valores_log["B3"],
        "paralelo": valores_log["B5"],
        "jornada": valores_log["B4"],
        "rectorDirector": valores_log["D3"],
        "tutorCurso": valores_log["D4"],
        "periodoExcel": valores_log["B6"],
    }

    campos_obligatorios = {
        "nombreInstitucion": "Nombre de la Institución (A1)",
        "anioLectivo": "Año Lectivo (D1/D2)",
        "codigoAmie": "Código AMIE (B7)",
        "gradoCurso": "Grado/Curso (B3)",
        "paralelo": "Paralelo (B5)",
        "jornada": "Jornada (B4)",
        "rectorDirector": "Rector/Director (D3)",
        "periodoExcel": "Período (B6)",
    }
    vacios = [descripcion for key, descripcion in campos_obligatorios.items() if not datos.get(key)]
    if vacios:
        raise ValueError("Formato incorrecto: celdas obligatorias vacías: " + ", ".join(vacios))

    registrar_debug("[extraer_datos_institucionales] Datos institucionales validados.")
    return datos


# 4. Inyección de datos en Plantillas HTML de Certificados
import re

PLANTILLAS_PERMITIDAS = {
    "FORMATO INICIAL 1.html",
    "FORMATO INICIAL 2.html",
    "PRIMERO DE EGB.html",
    "FORMALO DE ELEMENTAL.html",
    "FORMATO EGBM.html",
    "FORMATO EGBS.html",
    "FORMATO DE 1 Y 2 DE BGU.html",
    "FORMATO DE 3 DE BGU.html",
}

def _normalizar_grado(texto: str) -> str:
    """Normaliza el texto del grado para comparación robusta."""
    import unicodedata
    s = str(texto).upper().strip()
    # Remover tildes
    s = "".join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
    # Colapsar espacios múltiples
    s = " ".join(s.split())
    return s

def _es_egb(g: str) -> bool:
    """Verifica si el texto contiene indicador de EGB."""
    return any(k in g for k in ['EGB', 'EDUCACION GENERAL BASICA', 'EDUCACIÓN GENERAL BÁSICA'])

def _es_bachillerato(g: str) -> bool:
    """Verifica si el texto contiene indicador de Bachillerato."""
    return any(k in g for k in ['BACHILLERATO', 'BGU'])

def mapear_plantilla_python(grado_curso: str) -> str | None:
    """
    Mapea el grado/curso a la plantilla HTML correspondiente.
    ORDEN: Inicial → 3ro Bachillerato → 1ro/2do Bachillerato → 1ro EGB → EGB Elemental → EGB Media → EGB Superior → None
    """
    g = _normalizar_grado(grado_curso)
    if not g:
        return None

    # 1. INICIAL 1
    if 'INICIAL 1' in g or 'INICIAL I' in g and 'INICIAL 2' not in g and 'INICIAL II' not in g:
        return 'FORMATO INICIAL 1.html'

    # 2. INICIAL 2
    if 'INICIAL 2' in g or 'INICIAL II' in g:
        return 'FORMATO INICIAL 2.html'

    es_primero_bachillerato = any(k in g for k in ['PRIMERO', 'PRIMER', '1RO', '1ER', '1ERO'])
    es_segundo_bachillerato = any(k in g for k in ['SEGUNDO', '2DO'])
    es_tercero_bachillerato = any(k in g for k in ['TERCERO', '3RO', '3ER', 'TERCER'])

    # 3. 3RO DE BACHILLERATO
    if _es_bachillerato(g) and es_tercero_bachillerato:
        return 'FORMATO DE 3 DE BGU.html'

    # 4. 1RO y 2DO DE BACHILLERATO
    if _es_bachillerato(g) and (es_primero_bachillerato or es_segundo_bachillerato):
        return 'FORMATO DE 1 Y 2 DE BGU.html'

    # 5. 1RO DE EGB / PREPARATORIA
    if any(k in g for k in ['PRIMERO DE EGB', '1RO DE EGB', '1ERO DE EGB', '1RO EGB', '1ER GRADO', 'PREPARATORIA']):
        return 'PRIMERO DE EGB.html'

    # 6. 2DO, 3RO y 4TO DE EGB (ELEMENTAL) - requiere contexto EGB
    if 'ELEMENTAL' in g or (_es_egb(g) and any(k in g for k in ['SEGUNDO', 'TERCERO', 'TERCER', 'CUARTO', '2DO', '3RO', '4TO'])):
        return 'FORMALO DE ELEMENTAL.html'

    # 7. 5TO, 6TO y 7MO DE EGB (MEDIA) - requiere contexto EGB
    if 'MEDIA' in g or (_es_egb(g) and any(k in g for k in ['QUINTO', 'SEXTO', 'SEPTIMO', '5TO', '6TO', '7MO'])):
        return 'FORMATO EGBM.html'

    # 8. 8VO, 9NO y 10MO DE EGB (SUPERIOR) - requiere contexto EGB
    if 'SUPERIOR' in g or (_es_egb(g) and any(k in g for k in ['OCTAVO', 'NOVENO', 'DECIMO', '8VO', '9NO', '10MO'])):
        return 'FORMATO EGBS.html'

    # Sin coincidencia
    return None

def limpiar_celdas_dinamicas_html(html_content):
    """Vacía todos los campos académicos marcados antes de inyectar datos reales."""
    patron = re.compile(
        r'(<td\b[^>]*\bdata-academic-value=["\']true["\'][^>]*>)[\s\S]*?(</td>)',
        re.IGNORECASE,
    )
    return patron.sub(r"\g<1>\g<2>", html_content)


def inyectar_campo_certificado(html_content, campo, valor):
    """Inyecta texto escapado en un campo estable de datos personales o institucionales."""
    patron = re.compile(
        rf'(<(?P<tag>span|div|h1|p)\b[^>]*\bdata-cert-field=["\']{re.escape(campo)}["\'][^>]*>)'
        rf'[\s\S]*?(</(?P=tag)>)',
        re.IGNORECASE,
    )
    texto = escapar_html(mostrar_valor(valor))
    return patron.sub(lambda match: f"{match.group(1)}{texto}{match.group(3)}", html_content)


def inject_student_data_html(html_content, student, inst, logos):
    tutor_name = inst.get("tutor") or inst.get("tutorCurso") or ""
    rector_name = inst.get("rector") or inst.get("rectorDirector") or ""
    student_name_html = escapar_html(mostrar_valor(student.get("nombre", "")))
    student_id_html = escapar_html(mostrar_valor(student.get("cedula", "")))
    jornada_html = escapar_html(mostrar_valor(inst.get("jornada", "")))
    amie_html = escapar_html(mostrar_valor(inst.get("amie", "")))
    anio_html = escapar_html(mostrar_valor(inst.get("anio", "")))
    institucion_html = escapar_html(mostrar_valor(inst.get("nombre", "")))
    tutor_html = escapar_html(mostrar_valor(tutor_name))
    rector_html = escapar_html(mostrar_valor(rector_name))
    html_content = inyectar_campo_certificado(
        html_content, "institution-name", inst.get("nombre", "")
    )
    html_content = inyectar_campo_certificado(
        html_content, "amie", inst.get("amie", "")
    )
    html_content = inyectar_campo_certificado(
        html_content, "school-year", inst.get("anio", "")
    )
    html_content = inyectar_campo_certificado(
        html_content, "student-name", student.get("nombre", "")
    )
    html_content = inyectar_campo_certificado(
        html_content, "student-id", student.get("cedula", "")
    )
    grado_completo = (
        f"{inst.get('grado', '')} PARALELO: {inst.get('paralelo', '')}".strip()
    )
    grado_completo_html = escapar_html(grado_completo)
    html_content = inyectar_campo_certificado(
        html_content, "grade", grado_completo
    )
    html_content = inyectar_campo_certificado(
        html_content, "schedule", inst.get("jornada", "")
    )
    html_content = inyectar_campo_certificado(
        html_content, "tutor-name", tutor_name
    )
    html_content = inyectar_campo_certificado(
        html_content, "rector-name", rector_name
    )

    # 1. Reemplazar datos institucionales y del estudiante
    html_content = re.sub(
        r'(NOMBRE DEL ESTUDIANTE:\s*)[^<\n]+', 
        lambda match: f"{match.group(1)}{student_name_html}",
        html_content, 
        flags=re.IGNORECASE
    )
    html_content = re.sub(
        r'(CÉDULA:\s*)[^<\n]+', 
        lambda match: f"{match.group(1)}{student_id_html}",
        html_content, 
        flags=re.IGNORECASE
    )
    
    html_content = re.sub(
        r'(GRADO:\s*)[^<\n]+', 
        lambda match: f"{match.group(1)}{grado_completo_html}",
        html_content, 
        flags=re.IGNORECASE
    )
    
    html_content = re.sub(
        r'(JORNADA:\s*)[^<\n]+', 
        lambda match: f"{match.group(1)}{jornada_html}",
        html_content, 
        flags=re.IGNORECASE
    )
    
    html_content = re.sub(
        r'(AMIE:\s*)[^<\n\s,]+', 
        lambda match: f"{match.group(1)}{amie_html}",
        html_content, 
        flags=re.IGNORECASE
    )
    
    if 'data-cert-field="school-year"' not in html_content:
        html_content = re.sub(
            r'(AÑO LECTIVO:\s*)[^<\n]+',
            lambda match: f"{match.group(1)}{anio_html}",
            html_content,
            flags=re.IGNORECASE,
        )

    if inst.get("nombre"):
        html_content = re.sub(
            r'(<h1[^>]*>)[^<]*(EMILIANO HINOSTROZA|UNIDAD EDUCATIVA)[^<]*(</h1>)',
            lambda match: f"{match.group(1)}{institucion_html}{match.group(3)}",
            html_content,
            flags=re.IGNORECASE
        )
        
    # 2. Inyectar logos en base64
    if logos.get("logo1"):
        logo1_html = escapar_html(mostrar_valor(logos["logo1"]), quote=True)
        html_content = re.sub(
            r'id="img-logo-1"\s*class="[^"]*"',
            f'id="img-logo-1" class="absolute inset-0 w-full h-full object-contain p-1"',
            html_content
        )
        if 'src=' in html_content.split('id="img-logo-1"')[1].split('>')[0]:
            html_content = re.sub(
                r'id="img-logo-1"\s*src="[^"]*"',
                f'id="img-logo-1" src="{logo1_html}"',
                html_content
            )
        else:
            html_content = html_content.replace('id="img-logo-1"', f'id="img-logo-1" src="{logo1_html}"')
        html_content = html_content.replace('id="text-logo-1" class="', 'id="text-logo-1" class="hidden ')

    if logos.get("logo2"):
        logo2_html = escapar_html(mostrar_valor(logos["logo2"]), quote=True)
        html_content = re.sub(
            r'id="img-logo-2"\s*class="[^"]*"',
            f'id="img-logo-2" class="absolute inset-0 w-full h-full object-contain p-1"',
            html_content
        )
        if 'src=' in html_content.split('id="img-logo-2"')[1].split('>')[0]:
            html_content = re.sub(
                r'id="img-logo-2"\s*src="[^"]*"',
                f'id="img-logo-2" src="{logo2_html}"',
                html_content
            )
        else:
            html_content = html_content.replace('id="img-logo-2"', f'id="img-logo-2" src="{logo2_html}"')
        html_content = html_content.replace('id="text-logo-2" class="', 'id="text-logo-2" class="hidden ')

    # 3. Firmas
    if tutor_name:
        tutor_pattern = re.compile(r'<div[^>]*class="[^"]*text-center[^"]*"[\s\S]*?<div[^>]*class="[^"]*border-b[^"]*"[\s\S]*?</div>\s*<p[^>]*>\s*(TUTOR/A|TUTOR|TUTORA)\s*</p>\s*</div>', re.IGNORECASE)
        match = tutor_pattern.search(html_content)
        if match:
            block = match.group(0)
            new_block = block.replace(
                'class="border-b border-slate-800 mb-2 h-8 w-full"',
                'class="border-b border-slate-800 mb-2 w-full" style="height: 3.5rem;"'
            )
            if 'cert-nombre-firma' not in new_block:
                new_block = new_block.replace(
                    '<p>',
                    f'<p class="cert-nombre-firma font-bold text-slate-800 mt-1" style="font-size: 0.75rem;">{tutor_html}</p>\n            <p>'
                )
            html_content = html_content.replace(block, new_block)

    if rector_name:
        rector_pattern = re.compile(r'<div[^>]*class="[^"]*text-center[^"]*"[\s\S]*?<div[^>]*class="[^"]*border-b[^"]*"[\s\S]*?</div>\s*<p[^>]*>\s*(RECTOR/A|RECTOR|RECTORA|DIRECTOR|DIRECTORA)\s*</p>\s*</div>', re.IGNORECASE)
        match = rector_pattern.search(html_content)
        if match:
            block = match.group(0)
            new_block = block.replace(
                'class="border-b border-slate-800 mb-2 h-8 w-full"',
                'class="border-b border-slate-800 mb-2 w-full" style="height: 3.5rem;"'
            )
            if 'cert-nombre-firma' not in new_block:
                new_block = new_block.replace(
                    '<p>',
                    f'<p class="cert-nombre-firma font-bold text-slate-800 mt-1" style="font-size: 0.75rem;">{rector_html}</p>\n            <p>'
                )
            html_content = html_content.replace(block, new_block)

    return html_content

def inject_optativas_bgu3(html_content, materias_data):
    """Genera únicamente las optativas detectadas dentro del marcador estable de BGU_3."""
    patron = re.compile(
        r'(<tbody[^>]*\bid=["\']optativas-bgu3["\'][^>]*>)[\s\S]*?(</tbody>)',
        re.IGNORECASE,
    )
    if not patron.search(html_content):
        return html_content

    optativas = []
    for nombre, datos in materias_data.items():
        metadatos = completar_metadatos_asignatura(nombre, datos)
        if not metadatos.get("es_optativa_bgu3"):
            continue
        if metadatos.get("presentacion_certificado") != "escala_cualitativa":
            continue
        optativas.append((metadatos.get("orden", 999), metadatos["nombre"], datos))

    optativas.sort(key=lambda item: (int(item[0]), item[1]))
    filas = []
    for indice, (_, nombre, datos) in enumerate(optativas):
        clase_nombre = "text-left-cell font-semibold" + (" w-1/3" if indice == 0 else "")
        valores = [
            convertir_nota_optativa_a_escala_cualitativa(datos.get("t1")),
            convertir_nota_optativa_a_escala_cualitativa(datos.get("t2")),
            convertir_nota_optativa_a_escala_cualitativa(datos.get("t3")),
        ]
        filas.append(
            "\n                <tr>\n"
            f'                    <td class="{clase_nombre}">{escapar_html(nombre)}</td>\n'
            f"                    <td>{escapar_html(valores[0])}</td>\n"
            f"                    <td>{escapar_html(valores[1])}</td>\n"
            f"                    <td>{escapar_html(valores[2])}</td>\n"
            "                </tr>"
        )

    contenido = "".join(filas)
    return patron.sub(rf"\g<1>{contenido}\n            \g<2>", html_content, count=1)


def _inyectar_celda_academica(fila_html, campo, valor):
    patron = re.compile(
        rf'(<td\b[^>]*\bdata-academic-field=["\']{re.escape(campo)}["\'][^>]*>)'
        rf'[\s\S]*?(</td>)',
        re.IGNORECASE,
    )
    texto = escapar_html(mostrar_valor(valor))
    return patron.sub(lambda match: f"{match.group(1)}{texto}{match.group(2)}", fila_html)


def inject_subject_grades(html_content, materias_data, grado=None):
    html_content = inject_optativas_bgu3(html_content, materias_data)
    html_content = limpiar_celdas_dinamicas_html(html_content)

    materias_normalizadas = {
        normalizar_texto_asignatura(nombre): (nombre, datos)
        for nombre, datos in materias_data.items()
        if isinstance(datos, dict)
    }
    filas_marcadas = re.compile(
        r'<tr\b[^>]*\bdata-subject=["\'](?P<subject>[^"\']+)["\'][^>]*>'
        r'[\s\S]*?</tr>',
        re.IGNORECASE,
    )

    def inyectar_fila_marcada(match):
        fila = match.group(0)
        clave = normalizar_texto_asignatura(desescapar_html(match.group("subject")))
        encontrado = materias_normalizadas.get(clave)
        if not encontrado:
            return fila
        nombre, datos = encontrado
        metadatos = completar_metadatos_asignatura(nombre, datos)
        if metadatos.get("es_optativa_bgu3"):
            return fila

        es_cualitativa = metadatos["tipo"] == "cualitativa"
        valores = {
            "t1": presentar_nota_certificado(datos.get("t1"), grado, metadatos),
            "t2": presentar_nota_certificado(datos.get("t2"), grado, metadatos),
            "t3": presentar_nota_certificado(datos.get("t3"), grado, metadatos),
            "supletorio": (
                ""
                if es_cualitativa
                else presentar_nota_certificado(datos.get("supletorio"), grado, metadatos)
            ),
            "final": "",
        }
        if not es_cualitativa:
            valores["final"] = presentar_nota_certificado(
                datos.get("nota_final")
                if datos.get("nota_final") is not None
                else datos.get("promedio_anual"),
                grado,
                metadatos,
            )
        for campo, valor in valores.items():
            fila = _inyectar_celda_academica(fila, campo, valor)
        return fila

    html_content = filas_marcadas.sub(inyectar_fila_marcada, html_content)

    for sub_name, data in materias_data.items():
        metadatos = completar_metadatos_asignatura(sub_name, data)
        if metadatos.get("es_optativa_bgu3"):
            continue
        escaped_sub = re.escape(sub_name)
        tr_pattern = re.compile(
            rf'<tr[^>]*>\s*<td[^>]*>\s*{escaped_sub}[\s\S]*?</td>[\s\S]*?</tr>',
            re.IGNORECASE,
        )
        
        match = tr_pattern.search(html_content)
        if match:
            tr_block = match.group(0)
            if re.search(r"\bdata-subject=", tr_block, re.IGNORECASE):
                continue
            td_pattern = re.compile(r'<td[^>]*>([\s\S]*?)</td>', re.IGNORECASE)
            tds = list(td_pattern.finditer(tr_block))
            
            if len(tds) >= 2:
                grade_tds = tds[1:]
                es_cualitativa = data.get("tipo", tipo_asignatura(sub_name)) == "cualitativa"

                def fmt_cualitativo(valor):
                    return escapar_html(mostrar_valor(valor))

                def fmt(valor, _es_supletorio=False):
                    return escapar_html(
                        presentar_nota_certificado(valor, grado, metadatos)
                    )
                
                new_tds = []
                if es_cualitativa:
                    valores = [
                        fmt_cualitativo(data.get("t1")),
                        fmt_cualitativo(data.get("t2")),
                        fmt_cualitativo(data.get("t3")),
                    ]
                    for indice, td in enumerate(grade_tds):
                        apertura = re.match(r"<td[^>]*>", td.group(0), re.IGNORECASE).group(0)
                        valor = valores[indice] if indice < 3 else ""
                        new_tds.append(f"{apertura}{valor}</td>")
                if len(grade_tds) == 3:
                    if not es_cualitativa:
                        new_tds.append(f'<td>{fmt(data.get("t1"))}</td>')
                        new_tds.append(f'<td>{fmt(data.get("t2"))}</td>')
                        new_tds.append(f'<td>{fmt(data.get("t3"))}</td>')
                elif len(grade_tds) == 4:
                    if not es_cualitativa:
                        new_tds.append(f'<td>{fmt(data.get("t1"))}</td>')
                        new_tds.append(f'<td>{fmt(data.get("t2"))}</td>')
                        new_tds.append(f'<td>{fmt(data.get("t3"))}</td>')
                        val_final = data.get("nota_final") if data.get("nota_final") is not None else data.get("promedio_anual")
                        new_tds.append(f'<td class="font-bold bg-slate-50">{fmt(val_final)}</td>')
                elif len(grade_tds) == 5:
                    if not es_cualitativa:
                        new_tds.append(f'<td>{fmt(data.get("t1"))}</td>')
                        new_tds.append(f'<td>{fmt(data.get("t2"))}</td>')
                        new_tds.append(f'<td>{fmt(data.get("t3"))}</td>')
                        new_tds.append(f'<td>{fmt(data.get("supletorio"), True)}</td>')
                        val_final = data.get("nota_final") if data.get("nota_final") is not None else data.get("promedio_anual")
                        new_tds.append(f'<td class="font-bold bg-slate-50">{fmt(val_final)}</td>')
                    
                new_tr_block = tr_block[:tds[1].start()] + "".join(new_tds) + "</tr>"
                html_content = html_content.replace(tr_block, new_tr_block)
                
    # Reemplazar promedio general
    prom_pattern = re.compile(r'<tr[^>]*>[\s\S]*?<td[^>]*>\s*(PROMEDIO ANUAL|PROMEDIO GENERAL)\s*</td>[\s\S]*?</tr>', re.IGNORECASE)
    for m in prom_pattern.finditer(html_content):
        tr_block = m.group(0)
        td_pattern = re.compile(r'<td[^>]*>([\s\S]*?)</td>', re.IGNORECASE)
        tds = list(td_pattern.finditer(tr_block))
        if tds:
            try:
                grades = [
                    float(v.get("nota_final", 0.0) or v.get("promedio_anual", 0.0))
                    for nombre, v in materias_data.items()
                    if v.get("tipo", "cuantitativa") != "cualitativa"
                    and completar_metadatos_asignatura(nombre, v).get("permite_supletorio", True)
                    and (v.get("nota_final") is not None or v.get("promedio_anual") is not None)
                ]
                promedio_val = (
                    presentar_nota_certificado(
                        sum(grades) / len(grades),
                        grado,
                        {"tipo": "cuantitativa"},
                    )
                    if grades
                    else ""
                )
            except Exception:
                promedio_val = ""
                
            last_td = tds[-1]
            new_tr_block = tr_block[:last_td.start()] + f'<td class="font-bold bg-slate-100">{promedio_val}</td>' + "</tr>"
            html_content = html_content.replace(tr_block, new_tr_block)

    # Reemplazar cabecera "PROMEDIO ANUAL" a "Nota" en th
    html_content = re.sub(
        r'(<th[^>]*>\s*)PROMEDIO ANUAL(\s*</th>)',
        r'\g<1>Nota\g<2>',
        html_content,
        flags=re.IGNORECASE
    )
    return html_content


def inject_evaluacion_comportamental(html_content, evaluacion):
    """Inyecta T1/T2/T3 en la fila especial, sin tratarla como asignatura."""
    valores = evaluacion if isinstance(evaluacion, dict) else {}
    tr_pattern = re.compile(r"<tr[^>]*>[\s\S]*?</tr>", re.IGNORECASE)
    td_pattern = re.compile(r"<td[^>]*>[\s\S]*?</td>", re.IGNORECASE)

    for tr_match in tr_pattern.finditer(html_content):
        tr_block = tr_match.group(0)
        tds = list(td_pattern.finditer(tr_block))
        if len(tds) < 4:
            continue
        etiqueta = re.sub(r"<[^>]*>", " ", tds[0].group(0))
        if not es_encabezado_evaluacion_comportamental(etiqueta):
            continue

        celdas_nuevas = []
        for indice, periodo in enumerate(("T1", "T2", "T3"), start=1):
            celda_original = tds[indice].group(0)
            apertura = re.match(r"<td[^>]*>", celda_original, re.IGNORECASE).group(0)
            if "style=" not in apertura.lower():
                apertura = apertura[:-1] + ' style="white-space: normal; overflow-wrap: anywhere; vertical-align: top;">'
            celdas_nuevas.append(
                f"{apertura}{escapar_html(mostrar_valor(valores.get(periodo)))}</td>"
            )

        nuevo_bloque = (
            tr_block[:tds[1].start()]
            + "".join(celdas_nuevas)
            + tr_block[tds[3].end():]
        )
        return html_content.replace(tr_block, nuevo_bloque, 1)

    return html_content


def inject_asistencia_anual(
    html_content,
    asistencia,
    curso_id_esperado=None,
    estudiante_id_esperado=None,
):
    """Completa T1, T2, T3 y el anual sin recalcular sus resúmenes."""
    datos = asistencia if isinstance(asistencia, dict) else {}
    curso_id = mostrar_valor(datos.get("cursoId"))
    estudiante_id = mostrar_valor(datos.get("estudianteId"))
    curso_coincide = (
        not mostrar_valor(curso_id_esperado)
        or not curso_id
        or curso_id == mostrar_valor(curso_id_esperado)
    )
    estudiante_coincide = (
        not mostrar_valor(estudiante_id_esperado)
        or not estudiante_id
        or estudiante_id == mostrar_valor(estudiante_id_esperado)
    )
    identidad_coincide = curso_coincide and estudiante_coincide
    campos = {
        "registro": "totalFaltas",
        "justificacion": "justificadas",
        "injustificado": "injustificadas",
        "total": "totalAsistencia",
    }
    resumen_anual = datos.get("anual")
    if not isinstance(resumen_anual, dict):
        resumen_anual = datos

    periodos = {
        "T1": (datos.get("T1"), "configurado"),
        "T2": (datos.get("T2"), "configurado"),
        "T3": (datos.get("T3"), "configurado"),
        "ANUAL": (resumen_anual, "configurada"),
    }
    for periodo, (resumen, clave_configuracion) in periodos.items():
        resumen = resumen if isinstance(resumen, dict) else {}
        configurado = bool(
            identidad_coincide and resumen.get(clave_configuracion)
        )
        for campo, clave_dato in campos.items():
            valor = resumen.get(clave_dato) if configurado else ""
            patron = re.compile(
                rf'(<td\b'
                rf'(?=[^>]*\bdata-asistencia-periodo=["\']{periodo}["\'])'
                rf'(?=[^>]*\bdata-asistencia-campo=["\']{campo}["\'])'
                rf'[^>]*>)[\s\S]*?(</td>)',
                re.IGNORECASE,
            )
            texto = escapar_html(mostrar_valor(valor))
            html_content = patron.sub(rf"\g<1>{texto}\g<2>", html_content)

    # Compatibilidad con plantillas antiguas: reciben únicamente el total anual.
    if "data-asistencia-periodo" not in html_content.lower():
        configurada = bool(
            identidad_coincide and resumen_anual.get("configurada")
        )
        valores_anteriores = {
            "registro": resumen_anual.get("totalFaltas") if configurada else "",
            "justificadas": resumen_anual.get("justificadas") if configurada else "",
            "injustificadas": resumen_anual.get("injustificadas") if configurada else "",
            "total": resumen_anual.get("totalAsistencia") if configurada else "",
        }
        for clave, valor in valores_anteriores.items():
            patron = re.compile(
                rf'(<td[^>]*\bdata-asistencia=["\']{clave}["\'][^>]*>)[\s\S]*?(</td>)',
                re.IGNORECASE,
            )
            texto = escapar_html(mostrar_valor(valor))
            html_content = patron.sub(rf"\g<1>{texto}\g<2>", html_content)
    return html_content


def generar_certificados_inicial(payload):
    estudiantes = payload.get("datos_consolidados", [])
    inst = payload.get("institucion", {})
    logos = payload.get("logos", {})
    
    # Directorio de plantillas originales (solo lectura)
    templates_dir = os.path.join(os.path.dirname(__file__), "assets", "certificados")
    
    # Directorio de salida para certificados generados (recibido de Electron o fallback)
    cert_output_dir = os.path.abspath(
        payload.get(
            "certOutputDir",
            os.path.join(os.path.dirname(__file__), "assets", "certificados_generados"),
        )
    )
    os.makedirs(cert_output_dir, exist_ok=True)
    
    # Plantilla: JS tiene prioridad, luego fallback a mapear_plantilla_python
    plantilla_name_js = payload.get("plantillaName")
    grado_canonico = payload.get("gradoCursoCanonico", inst.get("grado", ""))
    
    # Validar plantilla del JS contra lista blanca
    if plantilla_name_js and plantilla_name_js in PLANTILLAS_PERMITIDAS:
        plantilla_name = plantilla_name_js
        registrar_debug("[Certificados-PY] Plantilla autorizada recibida.")
    else:
        plantilla_name = mapear_plantilla_python(grado_canonico)
        registrar_debug("[Certificados-PY] Plantilla determinada localmente.")
    
    if not plantilla_name:
        return {"error": f"No se encontró plantilla para el grado: {grado_canonico}", "supletorios": []}
    
    plantilla_path = os.path.join(templates_dir, plantilla_name)
    if not os.path.exists(plantilla_path):
        return {"error": f"Plantilla no encontrada en disco: {plantilla_name} → {plantilla_path}", "supletorios": []}
    
    registrar_debug(
        f"[Certificados-PY] Generación autorizada para {len(estudiantes)} estudiante(s)."
    )
    
    with open(plantilla_path, "r", encoding="utf-8") as f:
        template_html = f.read()
    
    supletorios_detectados = []
    archivos_generados = []
    
    for est in estudiantes:
        materias = {}
        for sub_name, datos_originales in est.get("materias", {}).items():
            if not isinstance(datos_originales, dict):
                registrar_debug(
                    "[Certificados-PY] Se omitió una estructura académica inesperada."
                )
                materias[sub_name] = {}
                continue
            materias[sub_name] = dict(datos_originales)

        for sub_name, m_data in list(materias.items()):
            metadatos = completar_metadatos_asignatura(sub_name, m_data)
            for campo in CAMPOS_METADATOS_ASIGNATURA:
                m_data[campo] = metadatos[campo]
            tipo = metadatos["tipo"]
            if tipo == "cualitativa":
                for periodo in ("t1", "t2", "t3"):
                    valor = m_data.get(periodo)
                    if valor is not None and not isinstance(valor, str):
                        registrar_debug(
                            "[Certificados-PY] Se conservó una valoración cualitativa no textual."
                        )
                m_data["t1"] = mostrar_valor(m_data.get("t1"))
                m_data["t2"] = mostrar_valor(m_data.get("t2"))
                m_data["t3"] = mostrar_valor(m_data.get("t3"))
                m_data["promedio_anual"] = None
                m_data["supletorio"] = None
                m_data["nota_final"] = None
                m_data["estado"] = None
                continue

            notas_trimestrales = [m_data.get("t1"), m_data.get("t2"), m_data.get("t3")]
            if any(not mostrar_valor(nota) for nota in notas_trimestrales):
                m_data["promedio_anual"] = None
                m_data["nota_final"] = None
                continue

            try:
                g1, g2, g3 = (float(nota) for nota in notas_trimestrales)
            except (TypeError, ValueError):
                # Una valoración ya cualitativa se conserva para presentación y no se
                # fuerza a participar en cálculos numéricos.
                continue
            p_anual = calcular_promedio_anual(g1, g2, g3)
            m_data["promedio_anual"] = p_anual
            if m_data.get("nota_final") is None:
                m_data["nota_final"] = p_anual
            if not metadatos["permite_supletorio"]:
                m_data["supletorio"] = None
                m_data["estado"] = None
                continue
                
            if 4.01 <= p_anual <= 6.99:
                nivel_str = inst.get("nivel", "")
                es_egb_media = "MEDIA" in str(inst.get("grado", "")).upper() or "MEDIA" in str(nivel_str).upper()
                supletorios_detectados.append({
                    "id": est["id_real"],
                    "nombre": est["nombre"],
                    "curso": inst.get("grado", ""),
                    "asignatura": sub_name,
                    "nivel": nivel_str,
                    "es_egb_media": es_egb_media
                })
        
        student_html = inject_student_data_html(template_html, est, inst, logos)
        student_html = inject_subject_grades(student_html, materias, grado_canonico)
        if curso_admite_evaluacion_comportamental(grado_canonico):
            student_html = inject_evaluacion_comportamental(
                student_html,
                est.get("evaluacion_comportamental", {}),
            )
        student_html = inject_asistencia_anual(
            student_html,
            est.get("asistencia"),
            payload.get("cursoActivoId"),
            est.get("id_real"),
        )
        
        # Guardar en directorio de salida, NO en plantillas
        student_id = str(est.get("id_real", "")).strip()
        if not es_segmento_ruta_seguro(student_id):
            raise ValueError("El identificador del estudiante no es válido.")
        output_filename = f"certificado_{student_id}.html"
        output_path = resolver_ruta_hija(cert_output_dir, output_filename)
        if output_path is None:
            raise ValueError("La ruta del certificado no es válida.")
        with output_path.open("w", encoding="utf-8") as f:
            f.write(student_html)
        
        archivos_generados.append(str(output_path))
            
    return {
        "supletorios": supletorios_detectados,
        "archivos": archivos_generados,
        "plantilla_usada": plantilla_name,
        "cert_output_dir": cert_output_dir
    }

def actualizar_certificado_supletorio(student_id, asignatura, nota_supletorio, cert_output_dir=None):
    """Actualiza un certificado HTML existente con la nota de supletorio."""
    if not es_segmento_ruta_seguro(student_id):
        return False
    metadatos = completar_metadatos_asignatura(asignatura)
    if metadatos["tipo"] == "cualitativa" or not metadatos["permite_supletorio"]:
        return False
    # Buscar en el directorio de salida proporcionado, o fallback
    if cert_output_dir:
        ruta_segura = resolver_ruta_hija(
            cert_output_dir,
            f"certificado_{student_id}.html",
        )
        if ruta_segura is None:
            return False
        file_path = str(ruta_segura)
    else:
        # Fallback: buscar en assets/certificados_generados y luego en assets/certificados (legacy)
        base_dir = os.path.dirname(__file__)
        file_path = os.path.join(base_dir, "assets", "certificados_generados", f"certificado_{student_id}.html")
        if not os.path.exists(file_path):
            file_path = os.path.join(base_dir, "assets", "certificados", f"certificado_{student_id}.html")
    
    if not os.path.exists(file_path):
        registrar_debug("[actualizar_certificado_supletorio] Certificado no localizado.")
        return False
        
    with open(file_path, "r", encoding="utf-8") as f:
        html_content = f.read()
        
    try:
        val_su = float(nota_supletorio)
    except (ValueError, TypeError):
        val_su = 0.0
        
    escaped_sub = re.escape(asignatura)
    tr_pattern = re.compile(rf'<tr[^>]*>\s*<td[^>]*>\s*{escaped_sub}\s*</td>[\s\S]*?</tr>', re.IGNORECASE)
    
    match = tr_pattern.search(html_content)
    if match:
        tr_block = match.group(0)
        td_pattern = re.compile(r'<td[^>]*>([\s\S]*?)</td>', re.IGNORECASE)
        tds = list(td_pattern.finditer(tr_block))
        
        # Helper para limpiar HTML y normalizar números
        clean_text = lambda html_str: re.sub(r'<[^>]*>', '', html_str).strip().replace(',', '.')
        
        if len(tds) == 6:
            try:
                t1_val = float(clean_text(tds[1].group(1)) or 0.0)
                t2_val = float(clean_text(tds[2].group(1)) or 0.0)
                t3_val = float(clean_text(tds[3].group(1)) or 0.0)
                prom_anual = truncar_2_decimales((t1_val + t2_val + t3_val) / 3.0)
            except Exception:
                prom_anual = 0.0

            val_final = 7.00 if val_su >= 7.00 else prom_anual
            
            supletorio_td = tds[4]
            new_tr_block = (
                tr_block[:supletorio_td.start()] +
                f'<td>{val_su:.2f}</td>' +
                f'<td class="font-bold bg-slate-50">{val_final:.2f}</td>' +
                '</tr>'
            )
            html_content = html_content.replace(tr_block, new_tr_block)
        elif len(tds) == 5:
            try:
                t1_val = float(clean_text(tds[1].group(1)) or 0.0)
                t2_val = float(clean_text(tds[2].group(1)) or 0.0)
                t3_val = float(clean_text(tds[3].group(1)) or 0.0)
                prom_anual = truncar_2_decimales((t1_val + t2_val + t3_val) / 3.0)
            except Exception:
                prom_anual = 0.0

            val_final = 7.00 if val_su >= 7.00 else prom_anual
            
            final_td = tds[4]
            new_tr_block = (
                tr_block[:final_td.start()] +
                f'<td class="font-bold bg-slate-50">{val_final:.2f}</td>' +
                '</tr>'
            )
            html_content = html_content.replace(tr_block, new_tr_block)
        else:
            return False
            
        with open(file_path, "w", encoding="utf-8") as f:
            f.write(html_content)
        return True
    return False


# 5. Proceso CLI Principal
def main():
    parser = argparse.ArgumentParser(description="Procesador de Notas Académicas de la UEEH")
    parser.add_argument('--analizar', action='store_true', help='Analiza los excels y devuelve la lista de estudiantes en JSON')
    parser.add_argument('--generar', action='store_true', help='Genera los PDF recibiendo el payload por stdin')
    parser.add_argument('--certificados', action='store_true', help='Genera los certificados HTML iniciales y detecta supletorios')
    parser.add_argument('--generar-formato-bgu3', action='store_true', help='Genera una copia XLSX de 3.º BGU con optativas')
    parser.add_argument(
        '--supletorios',
        nargs='?',
        const='-',
        help='Actualiza supletorios leyendo JSON desde stdin; mantiene compatibilidad con JSON como argumento.',
    )
    parser.add_argument('--t1', type=str, help='Ruta Excel Trimestre 1')
    parser.add_argument('--t2', type=str, help='Ruta Excel Trimestre 2')
    parser.add_argument('--t3', type=str, help='Ruta Excel Trimestre 3')
    parser.add_argument('--su', type=str, help='Ruta Excel Supletorio')
    parser.add_argument('--grado', type=str, help='Grado activo que autoriza las asignaturas del análisis')
    
    args = parser.parse_args()
    
    if args.generar_formato_bgu3:
        try:
            input_data = leer_stdin_utf8()
            if not input_data:
                raise ValueError("No se recibió la configuración del formato.")
            payload = json.loads(input_data)
            resultado = generar_formato_bgu3_con_optativas(
                payload.get("origen"),
                payload.get("destino"),
                payload.get("optativas"),
            )
            escribir_json_stdout(resultado)
        except Exception as error:
            escribir_json_stdout({
                "success": False,
                "error": str(error),
            })
        return

    if args.supletorios:
        try:
            contenido_updates = (
                leer_stdin_utf8()
                if args.supletorios == '-'
                else args.supletorios
            )
            updates = json.loads(contenido_updates)
            success_count = 0
            # Extract cert_output_dir from the first entry if provided
            cert_output_dir = None
            if isinstance(updates, list) and len(updates) > 0:
                cert_output_dir = updates[0].get("cert_output_dir")
            if isinstance(updates, list):
                for up in updates:
                    student_id = up.get("id")
                    asignatura = up.get("asignatura")
                    nota_su = up.get("nota_supletorio")
                    if student_id and asignatura and nota_su is not None:
                        if actualizar_certificado_supletorio(student_id, asignatura, nota_su, cert_output_dir):
                            success_count += 1
            escribir_json_stdout({"success": True, "updated": success_count})
        except Exception:
            escribir_json_stdout({
                "success": False,
                "error": "No se pudieron validar o actualizar las notas de supletorio.",
            })
        return

    if args.certificados:
        try:
            # Leer el payload JSON enviado desde Electron a través de stdin
            input_data = leer_stdin_utf8()
            if not input_data:
                escribir_json_stdout({"success": False, "error": "No se recibió payload en stdin"})
                return
                
            payload = json.loads(input_data)
            result = generar_certificados_inicial(payload)
            
            # result is now a dict with supletorios, archivos, plantilla_usada, cert_output_dir
            if isinstance(result, dict) and "error" in result:
                escribir_json_stdout({"success": False, "error": result["error"]})
            else:
                escribir_json_stdout({"success": True, **result})
        except Exception:
            escribir_json_stdout({
                "success": False,
                "error": "No se pudieron generar los certificados.",
            })
        return

    elif args.analizar:
        try:
            if not args.grado or not args.grado.strip():
                raise ValueError("No se recibió el grado del curso seleccionado.")
            first_path = args.t1 or args.t2 or args.t3 or args.su
            datos_inst = extraer_datos_institucionales(first_path)
            diagnostico = crear_diagnostico_asignaturas()
            
            # Cargar datos del excel específico
            records = cargar_excel_datos(first_path, args.grado, diagnostico)
            
            # Obtener asignaturas
            asignaturas = sorted(
                dict.fromkeys(item["canonica"] for item in diagnostico["asignaturasReconocidas"]),
                key=orden_asignatura,
            )
            if not asignaturas:
                partes_error = [f"No se reconoció ninguna asignatura válida para {args.grado}."]
                ignoradas = [item["original"] for item in diagnostico["asignaturasIgnoradasPorCurso"]]
                desconocidas = [item["original"] for item in diagnostico["asignaturasNoReconocidas"]]
                if ignoradas:
                    partes_error.append("No corresponden al curso: " + ", ".join(ignoradas) + ".")
                if desconocidas:
                    partes_error.append("No están en el catálogo oficial: " + ", ".join(desconocidas) + ".")
                escribir_json_stdout({
                    "error": " ".join(partes_error),
                    **diagnostico,
                })
                return
                
            estudiantes = list(records.values())
            
            # Calcular promedios para la tabla
            estudiantes_tabla = []
            for est in estudiantes:
                grades_list = [
                    valor
                    for nombre, valor in est["notas"].items()
                    if isinstance(valor, (int, float))
                    and est.get("metadatos_asignaturas", {})
                        .get(nombre, metadatos_asignatura(nombre))
                        .get("permite_supletorio", True)
                ]
                prom = sum(grades_list) / len(grades_list) if grades_list else None
                estudiantes_tabla.append({
                    "cedula": est["cedula"],
                    "nombre": est["nombre"],
                    "promedio": truncar_2_decimales(prom) if prom is not None else None,
                    "estado": "APROBADO" if prom is not None and prom >= 7.0 else (
                        "SUPLETORIO" if prom is not None else "PENDIENTE DE NOTAS EXTERNAS"
                    ),
                })
            
            output_data = {
                "datosInstitucion": datos_inst,
                "periodoExcel": datos_inst.get("periodoExcel", ""),
                "asignaturas": asignaturas,
                "estudiantes": estudiantes,
                "estudiantes_tabla": estudiantes_tabla,
                **diagnostico,
            }
            escribir_json_stdout(output_data)
        except Exception as e:
            escribir_json_stdout({"error": str(e)})
            
    elif args.generar:
        try:
            # Leer el payload JSON enviado desde Electron a través de stdin
            input_data = leer_stdin_utf8()
            if not input_data:
                escribir_json_stdout({"success": False, "error": "No se recibió payload en stdin"})
                return
                
            payload = json.loads(input_data)
            
            excels = payload.get("excels", {})
            logos = payload.get("logos", {})
            inst = payload.get("institucion", {})
            estudiantes_seleccionados = payload.get("estudiantes", [])
            
            # Consolidar todos los estudiantes de los excels
            if payload.get("datos_consolidados"):
                seleccionados_data = payload.get("datos_consolidados")
            else:
                datos_completos = consolidar_estudiantes(
                    excels.get("t1"), excels.get("t2"), excels.get("t3"), excels.get("su"), inst.get("grado")
                )
                seleccionados_data = [est for est in datos_completos if est["id_real"] in estudiantes_seleccionados]
            
            if not seleccionados_data:
                escribir_json_stdout({"success": False, "error": "No se encontraron datos para los estudiantes seleccionados"})
                return
                
            output_pdf = payload.get("outputPath")
            if (
                not isinstance(output_pdf, str)
                or not os.path.isabs(output_pdf)
                or Path(output_pdf).suffix.lower() != ".pdf"
                or not Path(output_pdf).parent.is_dir()
            ):
                escribir_json_stdout({
                    "success": False,
                    "error": "No se recibió un destino PDF autorizado.",
                })
                return
            output_pdf = str(Path(output_pdf).resolve())
            
            generar_boletin_pdf(seleccionados_data, inst, logos, output_pdf)
            
            escribir_json_stdout({"success": True, "path": output_pdf})
            
        except Exception:
            escribir_json_stdout({
                "success": False,
                "error": "No se pudo generar el PDF solicitado.",
            })
    else:
        parser.print_help()

if __name__ == "__main__":
    main()
