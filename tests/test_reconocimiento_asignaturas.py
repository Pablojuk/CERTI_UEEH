# -*- coding: utf-8 -*-

import json
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook

from catalogo_asignaturas import (
    CATALOGO_ASIGNATURAS,
    clasificar_asignatura,
    grados_equivalentes,
    metadatos_asignatura,
    normalizar_texto_asignatura,
    tipo_asignatura,
)
from procesador_notas import (
    ErrorGradoExcel,
    actualizar_certificado_supletorio,
    cargar_excel_datos,
    consolidar_estudiantes,
    convertir_nota_optativa_a_escala_cualitativa,
    crear_diagnostico_asignaturas,
    curso_admite_evaluacion_comportamental,
    generar_certificados_inicial,
    inject_evaluacion_comportamental,
    inject_optativas_bgu3,
    inject_subject_grades,
    truncar_2_decimales,
)


OPTATIVAS_BGU3 = [
    "APRECIACIÓN MUSICAL",
    "CORRIENTES FILOSÓFICAS",
    "INVESTIGACIÓN EN CIENCIA Y TECNOLOGÍA",
    "LECTURA CRÍTICA DE MENSAJES",
    "LENGUA KICHWA",
    "REDACCIÓN CREATIVA",
    "SOCIOLOGÍA",
    "TEATRO",
    "PROBLEMAS DEL MUNDO CONTEMPORÁNEO",
    "PSICOLOGÍA",
    "LENGUA EXTRANJERA: FRANCÉS",
    "FÍSICA SUPERIOR",
    "QUÍMICA SUPERIOR",
    "MATEMÁTICA SUPERIOR",
    "BIOLOGÍA SUPERIOR",
    "LENGUA Y CULTURA ANCESTRAL",
]


class ReconocimientoAsignaturasTests(unittest.TestCase):
    def setUp(self):
        self.temporal = tempfile.TemporaryDirectory()
        self.directorio = Path(self.temporal.name)

    def tearDown(self):
        self.temporal.cleanup()

    def crear_excel(self, nombre, grado, materias, cedula="0123456789", nombre_estudiante="ESTUDIANTE PRUEBA"):
        ruta = self.directorio / nombre
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Periodo"
        ws["A1"] = "UNIDAD EDUCATIVA PRUEBA"
        ws["D1"] = "2026-2027"
        ws["B3"] = grado
        ws["D3"] = "RECTOR PRUEBA"
        ws["B4"] = "MATUTINA"
        ws["D4"] = "TUTOR PRUEBA"
        ws["B5"] = "A"
        ws["B6"] = "TRIMESTRE 1"
        ws["B7"] = "01H00000"
        ws.cell(row=9, column=1, value="LISTADO")
        ws.cell(row=9, column=2, value="CEDULA")
        ws.cell(row=10, column=1, value=nombre_estudiante)
        ws.cell(row=10, column=2, value=cedula)
        for indice, (materia, nota) in enumerate(materias, start=3):
            ws.cell(row=9, column=indice, value=materia)
            ws.cell(row=10, column=indice, value=nota)
        wb.save(ruta)
        return ruta

    def crear_excel_civica_real_controlado(self, nombre, valores):
        ruta = self.directorio / nombre
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Periodo"
        ws["A1"] = "UNIDAD EDUCATIVA PRUEBA"
        ws["D1"] = "2026-2027"
        ws["B3"] = "3RO BGU"
        ws["D3"] = "RECTOR PRUEBA"
        ws["B4"] = "MATUTINA"
        ws["D4"] = "TUTOR PRUEBA"
        ws["B5"] = "A"
        ws["B6"] = "TRIMESTRE 1"
        ws["B7"] = "01H00000"
        ws["A9"] = "LISTADO"
        ws["B9"] = "CEDULA"
        ws["M9"] = "CÍVICA Y ACOMPAÑAMIENTO INTEGRAL EN EL AULA"
        for indice, valor in enumerate(valores, start=10):
            ws.cell(indice, 1, f"ESTUDIANTE {indice - 9}")
            ws.cell(indice, 2, f"01234567{indice - 10:02d}")
            ws.cell(indice, 13, valor)
        wb.save(ruta)
        return ruta

    def agregar_comportamiento(self, ruta, valor, encabezado="EVALUACIÓN COMPORTAMENTAL", columna=18):
        from openpyxl import load_workbook
        wb = load_workbook(ruta)
        ws = wb["Reporte Periodo"]
        ws.cell(row=9, column=columna, value=encabezado)
        ws.cell(row=10, column=columna, value=valor)
        wb.save(ruta)

    def contenido_optativas(self, html):
        import re
        coincidencia = re.search(
            r'<tbody[^>]*id="optativas-bgu3"[^>]*>([\s\S]*?)</tbody>',
            html,
            re.IGNORECASE,
        )
        self.assertIsNotNone(coincidencia)
        return coincidencia.group(1)

    def test_normaliza_tildes_espacios_mayusculas_y_separadores(self):
        self.assertEqual(normalizar_texto_asignatura("  Biología --  "), "BIOLOGIA")
        self.assertEqual(normalizar_texto_asignatura("educación / física"), "EDUCACION FISICA")

    def test_truncamiento_no_resta_una_centesima_por_punto_flotante(self):
        self.assertEqual(truncar_2_decimales(9.2), 9.2)

    def test_catalogo_declara_civica_cualitativa_y_mantiene_numericas(self):
        self.assertEqual(tipo_asignatura("CÍVICA Y ACOMPAÑAMIENTO INTEGRAL EN EL AULA"), "cualitativa")
        for materia in ("MATEMÁTICA", "BIOLOGÍA", "LENGUA Y LITERATURA"):
            with self.subTest(materia=materia):
                self.assertEqual(tipo_asignatura(materia), "cuantitativa")

    def test_archivo_controlado_reproduce_m10_m16_sin_alterar_texto(self):
        esperados = ["B+", "B+", "A-", "b+", "B+", "b+", "B+"]
        ruta = self.crear_excel_civica_real_controlado("t1_controlado.xlsx", esperados)
        estudiantes = list(cargar_excel_datos(ruta, "3RO BGU").values())
        obtenidos = [
            estudiante["notas"]["CÍVICA Y ACOMPAÑAMIENTO INTEGRAL EN EL AULA"]
            for estudiante in estudiantes
        ]
        self.assertEqual(obtenidos, esperados)
        self.assertTrue(all(
            estudiante["tipos_asignaturas"]["CÍVICA Y ACOMPAÑAMIENTO INTEGRAL EN EL AULA"] == "cualitativa"
            for estudiante in estudiantes
        ))

    def test_cli_analizar_conserva_tipo_y_valores_cualitativos(self):
        esperados = ["B+", "B+", "A-", "b+", "B+", "b+", "B+"]
        ruta = self.crear_excel_civica_real_controlado("t1_cli_controlado.xlsx", esperados)
        proceso = subprocess.run(
            [
                sys.executable,
                str(Path(__file__).parents[1] / "procesador_notas.py"),
                "--analizar",
                "--grado",
                "3RO BGU",
                "--t1",
                str(ruta),
            ],
            capture_output=True,
            text=True,
            encoding="utf-8",
            check=True,
        )
        respuesta = json.loads(proceso.stdout)
        self.assertEqual(
            [
                estudiante["notas"]["CÍVICA Y ACOMPAÑAMIENTO INTEGRAL EN EL AULA"]
                for estudiante in respuesta["estudiantes"]
            ],
            esperados,
        )
        self.assertTrue(all(
            estudiante["tipos_asignaturas"]["CÍVICA Y ACOMPAÑAMIENTO INTEGRAL EN EL AULA"] == "cualitativa"
            for estudiante in respuesta["estudiantes"]
        ))
        self.assertTrue(all(estudiante["promedio"] is None for estudiante in respuesta["estudiantes_tabla"]))
        self.assertTrue(all(
            estudiante["estado"] == "PENDIENTE DE NOTAS EXTERNAS"
            for estudiante in respuesta["estudiantes_tabla"]
        ))

    def test_civica_solo_t1_conserva_texto_y_deja_t2_t3_vacios(self):
        ruta = self.crear_excel_civica_real_controlado("solo_t1_civica.xlsx", ["b+"])
        materia = consolidar_estudiantes(ruta, None, None, None, "3RO BGU")[0]["materias"][
            "CÍVICA Y ACOMPAÑAMIENTO INTEGRAL EN EL AULA"
        ]
        self.assertEqual(materia["tipo"], "cualitativa")
        self.assertEqual((materia["t1"], materia["t2"], materia["t3"]), ("b+", "", ""))
        self.assertIsNone(materia["promedio_anual"])
        self.assertIsNone(materia["nota_final"])
        self.assertIsNone(materia["estado"])
        plantilla = Path(__file__).parents[1] / "assets" / "certificados" / "FORMATO DE 3 DE BGU.html"
        html = inject_subject_grades(plantilla.read_text(encoding="utf-8"), {
            "CÍVICA Y ACOMPAÑAMIENTO INTEGRAL EN EL AULA": materia
        })
        indice = html.index("CÍVICA Y ACOMPAÑAMIENTO INTEGRAL EN EL AULA")
        fila = html[html.rfind("<tr", 0, indice):html.index("</tr>", indice) + len("</tr>")]
        self.assertRegex(fila, r"<td>b\+</td><td></td><td></td>")

    def test_valor_cualitativo_vacio_permanece_ausente_sin_marcadores(self):
        ruta = self.crear_excel("civica_vacia.xlsx", "3RO BGU", [
            ("CÍVICA Y ACOMPAÑAMIENTO INTEGRAL EN EL AULA", None),
            ("MATEMÁTICA", 8),
        ])
        registro = next(iter(cargar_excel_datos(ruta, "3RO BGU").values()))
        self.assertIsNone(registro["notas"]["CÍVICA Y ACOMPAÑAMIENTO INTEGRAL EN EL AULA"])
        html = inject_subject_grades(
            (Path(__file__).parents[1] / "assets" / "certificados" / "FORMATO DE 3 DE BGU.html").read_text(encoding="utf-8"),
            {
                "CÍVICA Y ACOMPAÑAMIENTO INTEGRAL EN EL AULA": {
                    "tipo": "cualitativa", "t1": None, "t2": "", "t3": None
                }
            },
        )
        indice = html.index("CÍVICA Y ACOMPAÑAMIENTO INTEGRAL EN EL AULA")
        fila = html[html.rfind("<tr", 0, indice):html.index("</tr>", indice) + len("</tr>")]
        self.assertNotRegex(fila, r">\s*(0|0\.00|None|nan|PEND(?:IENTE)?|undefined)\s*</td>")

    def test_civica_con_tres_trimestres_no_genera_promedio_estado_ni_supletorio(self):
        rutas = [
            self.crear_excel(f"civica_t{indice}.xlsx", "3RO BGU", [
                ("CÍVICA Y ACOMPAÑAMIENTO INTEGRAL EN EL AULA", valor)
            ])
            for indice, valor in enumerate(("B+", "A-", "B"), start=1)
        ]
        materia = consolidar_estudiantes(*rutas, None, "3RO BGU")[0]["materias"][
            "CÍVICA Y ACOMPAÑAMIENTO INTEGRAL EN EL AULA"
        ]
        self.assertEqual((materia["t1"], materia["t2"], materia["t3"]), ("B+", "A-", "B"))
        for campo in ("promedio_anual", "supletorio", "nota_final", "estado"):
            self.assertIsNone(materia[campo])

    def test_civica_no_modifica_promedio_general_cuantitativo(self):
        rutas = [
            self.crear_excel(f"prom_civica_t{indice}.xlsx", "3RO BGU", [
                ("MATEMÁTICA", 9),
                ("CÍVICA Y ACOMPAÑAMIENTO INTEGRAL EN EL AULA", valor),
            ])
            for indice, valor in enumerate(("B+", "A-", "b+"), start=1)
        ]
        estudiante = consolidar_estudiantes(*rutas, None, "3RO BGU")[0]
        self.assertEqual(estudiante["promedio"], 9)
        self.assertEqual(estudiante["estado"], "APROBADO")

    def test_logica_cualitativa_es_general_y_no_depende_del_nombre_civica(self):
        clasificacion = {
            "estado": "reconocida",
            "original": "TALLER FORMATIVO",
            "canonica": "TALLER FORMATIVO",
            "metodo": "coincidencia_exacta",
            "tipo": "cualitativa",
        }
        ruta = self.crear_excel("cualitativa_generica.xlsx", "3RO BGU", [("TALLER FORMATIVO", "Excelente-")])
        with patch("procesador_notas.clasificar_asignatura", return_value=clasificacion):
            estudiante = next(iter(cargar_excel_datos(ruta, "3RO BGU").values()))
        self.assertEqual(estudiante["notas"]["TALLER FORMATIVO"], "Excelente-")
        self.assertEqual(estudiante["tipos_asignaturas"]["TALLER FORMATIVO"], "cualitativa")

    def test_inyeccion_html_cualitativa_es_literal_y_escapada(self):
        plantilla = Path(__file__).parents[1] / "assets" / "certificados" / "FORMATO DE 3 DE BGU.html"
        html = inject_subject_grades(
            plantilla.read_text(encoding="utf-8"),
            {
                "CÍVICA Y ACOMPAÑAMIENTO INTEGRAL EN EL AULA": {
                    "tipo": "cualitativa",
                    "t1": "b+",
                    "t2": "A-",
                    "t3": "<B+>",
                    "promedio_anual": None,
                    "supletorio": None,
                    "nota_final": None,
                }
            },
        )
        self.assertIn(">b+</td>", html)
        self.assertIn(">A-</td>", html)
        self.assertIn("&lt;B+&gt;", html)
        self.assertNotIn(">0.00</td>", html)

    def test_inyeccion_cualitativa_funciona_en_todas_las_plantillas_con_civica(self):
        directorio = Path(__file__).parents[1] / "assets" / "certificados"
        plantillas = [
            ruta for ruta in directorio.glob("*.html")
            if "CÍVICA Y ACOMPAÑAMIENTO INTEGRAL EN EL AULA" in ruta.read_text(encoding="utf-8")
        ]
        self.assertGreaterEqual(len(plantillas), 5)
        for plantilla in plantillas:
            with self.subTest(plantilla=plantilla.name):
                html = inject_subject_grades(
                    plantilla.read_text(encoding="utf-8"),
                    {
                        "CÍVICA Y ACOMPAÑAMIENTO INTEGRAL EN EL AULA": {
                            "tipo": "cualitativa", "t1": "B+", "t2": "A-", "t3": "b+"
                        }
                    },
                )
                indice = html.index("CÍVICA Y ACOMPAÑAMIENTO INTEGRAL EN EL AULA")
                fila = html[html.rfind("<tr", 0, indice):html.index("</tr>", indice) + len("</tr>")]
                self.assertIn(">B+</td>", fila)
                self.assertIn(">A-</td>", fila)
                self.assertIn(">b+</td>", fila)

    def test_certificado_no_detecta_supletorio_para_cualitativas(self):
        resultado = generar_certificados_inicial({
            "datos_consolidados": [{
                "id_real": "0123456789",
                "cedula": "0123456789",
                "nombre": "ESTUDIANTE PRUEBA",
                "materias": {
                    "MATEMÁTICA": {"tipo": "cuantitativa", "t1": 6, "t2": 6, "t3": 6},
                    "CÍVICA Y ACOMPAÑAMIENTO INTEGRAL EN EL AULA": {
                        "tipo": "cualitativa", "t1": "B+", "t2": "A-", "t3": "b+"
                    },
                },
                "evaluacion_comportamental": {},
            }],
            "institucion": {"grado": "3RO BGU", "nivel": "BGU"},
            "logos": {},
            "gradoCursoCanonico": "3RO BGU",
            "plantillaName": "FORMATO DE 3 DE BGU.html",
            "certOutputDir": str(self.directorio / "certificados"),
        })
        self.assertEqual([item["asignatura"] for item in resultado["supletorios"]], ["MATEMÁTICA"])
        html = Path(resultado["archivos"][0]).read_text(encoding="utf-8")
        self.assertIn(">B+</td>", html)
        self.assertIn(">A-</td>", html)
        self.assertIn(">b+</td>", html)

    def test_bologia_se_reconoce_como_biologia(self):
        resultado = clasificar_asignatura("Bologia", "1RO BGU")
        self.assertEqual(resultado["canonica"], "BIOLOGÍA")
        self.assertEqual(resultado["metodo"], "coincidencia_aproximada")

    def test_biolgia_se_reconoce_como_biologia(self):
        self.assertEqual(clasificar_asignatura("Biolgia", "1RO BGU")["canonica"], "BIOLOGÍA")

    def test_matemtica_se_reconoce_como_matematica(self):
        self.assertEqual(clasificar_asignatura("Matemtica", "1RO BGU")["canonica"], "MATEMÁTICA")

    def test_matematica_superior_se_ignora_en_primero_bgu(self):
        resultado = clasificar_asignatura("MATEMÁTICA SUPERIOR", "1RO BGU")
        self.assertEqual(resultado["estado"], "ignorada_por_curso")
        self.assertEqual(resultado["canonica"], "MATEMÁTICA SUPERIOR")

    def test_matematica_superior_se_admite_en_tercero_bgu(self):
        resultado = clasificar_asignatura("MATEMÁTICA SUPERIOR", "3RO BGU")
        self.assertEqual(resultado["estado"], "reconocida")
        self.assertEqual(resultado["canonica"], "MATEMÁTICA SUPERIOR")

    def test_matematica_se_admite_en_primero_bgu(self):
        resultado = clasificar_asignatura("MATEMÁTICA", "1RO BGU")
        self.assertEqual(resultado["estado"], "reconocida")

    def test_matematica_superior_nunca_se_convierte_en_matematica(self):
        for texto in ("MATEMÁTICA SUPERIOR", "MATEMATICA SUPRIOR"):
            for grado in ("1RO BGU", "3RO BGU"):
                self.assertEqual(clasificar_asignatura(texto, grado)["canonica"], "MATEMÁTICA SUPERIOR")

    def test_quimica_superior_nunca_se_convierte_en_quimica(self):
        for texto in ("QUÍMICA SUPERIOR", "QUIMICA SUPERIO"):
            resultado = clasificar_asignatura(texto, "1RO BGU")
            self.assertEqual(resultado["estado"], "ignorada_por_curso")
            self.assertEqual(resultado["canonica"], "QUÍMICA SUPERIOR")

    def test_variantes_de_tres_trimestres_se_consolidan(self):
        t1 = self.crear_excel("t1.xlsx", "1RO BGU", [("BIOLOGÍA", 8.25)])
        t2 = self.crear_excel("t2.xlsx", "PRIMERO DE BACHILLERATO", [("Bologia", 8.5)])
        t3 = self.crear_excel("t3.xlsx", "PRIMER BGU", [("Biologia", 8.75)])
        resultado = consolidar_estudiantes(t1, t2, t3, None, "1RO DE BACHILLERATO CIENCIAS")
        self.assertEqual(list(resultado[0]["materias"]), ["BIOLOGÍA"])
        self.assertEqual(resultado[0]["materias"]["BIOLOGÍA"]["t1"], 8.25)
        self.assertEqual(resultado[0]["materias"]["BIOLOGÍA"]["t2"], 8.5)
        self.assertEqual(resultado[0]["materias"]["BIOLOGÍA"]["t3"], 8.75)

    def test_limita_encabezados_a_columna_x_y_conserva_comportamiento(self):
        ruta = self.crear_excel("limite_x.xlsx", "3RO BGU", [("MATEMÁTICA", 9)])
        self.agregar_comportamiento(
            ruta,
            "Transforma los desacuerdos en oportunidades de crecimiento y cooperación",
            encabezado="  Evaluación   comportamental  ",
        )
        from openpyxl import load_workbook
        wb = load_workbook(ruta)
        ws = wb["Reporte Periodo"]
        for columna, encabezado in ((25, "EDUCACION INICIAL"), (26, "4TO DE EGB"), (27, "F")):
            ws.cell(row=9, column=columna, value=encabezado)
            ws.cell(row=10, column=columna, value=7)
        wb.save(ruta)

        diagnostico = crear_diagnostico_asignaturas()
        estudiante = next(iter(cargar_excel_datos(ruta, "3RO BGU", diagnostico).values()))
        self.assertEqual(estudiante["notas"], {"MATEMÁTICA": 9.0})
        self.assertEqual(
            estudiante["evaluacion_comportamental"],
            "Transforma los desacuerdos en oportunidades de crecimiento y cooperación",
        )
        mensajes = json.dumps(diagnostico, ensure_ascii=False)
        self.assertNotIn("EDUCACION INICIAL", mensajes)
        self.assertNotIn("4TO DE EGB", mensajes)
        self.assertNotIn('"F"', mensajes)

    def test_encabezado_comportamental_acepta_tildes_mayusculas_y_espacios(self):
        variantes = (
            "EVALUACIÓN COMPORTAMENTAL",
            "EVALUACION COMPORTAMENTAL",
            "Evaluación comportamental",
            "  EVALUACIÓN  COMPORTAMENTAL  ",
        )
        for indice, encabezado in enumerate(variantes):
            with self.subTest(encabezado=encabezado):
                ruta = self.crear_excel(f"comportamiento_{indice}.xlsx", "3RO BGU", [("MATEMÁTICA", 9)])
                self.agregar_comportamiento(ruta, f"Texto {indice}", encabezado)
                diagnostico = crear_diagnostico_asignaturas()
                estudiante = next(iter(cargar_excel_datos(ruta, "3RO BGU", diagnostico).values()))
                self.assertEqual(estudiante["evaluacion_comportamental"], f"Texto {indice}")
                self.assertEqual(list(estudiante["notas"]), ["MATEMÁTICA"])
                self.assertNotIn(encabezado.strip(), json.dumps(diagnostico, ensure_ascii=False))

    def test_comportamiento_de_cada_trimestre_se_consolida_sin_copias(self):
        rutas = []
        for periodo, texto in enumerate(("Comportamiento T1", "Comportamiento T2", "Comportamiento T3"), start=1):
            ruta = self.crear_excel(f"comp_t{periodo}.xlsx", "3RO BGU", [("MATEMÁTICA", 9)])
            self.agregar_comportamiento(ruta, texto)
            rutas.append(ruta)
        estudiante = consolidar_estudiantes(*rutas, None, "3RO BGU")[0]
        self.assertEqual(
            estudiante["evaluacion_comportamental"],
            {"T1": "Comportamiento T1", "T2": "Comportamiento T2", "T3": "Comportamiento T3"},
        )

    def test_solo_t1_no_rellena_comportamiento_de_t2_ni_t3(self):
        t1 = self.crear_excel("solo_comp_t1.xlsx", "3RO BGU", [("MATEMÁTICA", 9)])
        self.agregar_comportamiento(t1, "Solo comportamiento T1")
        estudiante = consolidar_estudiantes(t1, None, None, None, "3RO BGU")[0]
        self.assertEqual(
            estudiante["evaluacion_comportamental"],
            {"T1": "Solo comportamiento T1", "T2": "", "T3": ""},
        )

    def test_comportamiento_vacio_permanece_vacio(self):
        rutas = []
        for periodo, texto in enumerate(("Comportamiento T1", None, "Comportamiento T3"), start=1):
            ruta = self.crear_excel(f"comp_vacio_t{periodo}.xlsx", "3RO BGU", [("MATEMÁTICA", 9)])
            self.agregar_comportamiento(ruta, texto)
            rutas.append(ruta)
        estudiante = consolidar_estudiantes(*rutas, None, "3RO BGU")[0]
        self.assertEqual(estudiante["evaluacion_comportamental"]["T2"], "")
        self.assertNotEqual(estudiante["evaluacion_comportamental"]["T1"], estudiante["evaluacion_comportamental"]["T2"])

    def test_nota_vacia_no_se_convierte_en_cero(self):
        t1 = self.crear_excel("nota_vacia_t1.xlsx", "3RO BGU", [("MATEMÁTICA", None), ("BIOLOGÍA", 9)])
        t2 = self.crear_excel("nota_vacia_t2.xlsx", "3RO BGU", [("MATEMÁTICA", 8.5), ("BIOLOGÍA", 9.2)])
        t3 = self.crear_excel("nota_vacia_t3.xlsx", "3RO BGU", [("MATEMÁTICA", 8.8), ("BIOLOGÍA", 9.4)])
        estudiante = consolidar_estudiantes(t1, t2, t3, None, "3RO BGU")[0]
        self.assertIn("MATEMÁTICA", estudiante["materias"])
        self.assertIsNone(estudiante["materias"]["MATEMÁTICA"]["t1"])
        self.assertEqual(estudiante["materias"]["MATEMÁTICA"]["t2"], 8.5)
        self.assertEqual(estudiante["materias"]["BIOLOGÍA"]["t1"], 9.0)
        self.assertIsNone(estudiante["materias"]["MATEMÁTICA"]["promedio_anual"])

    def test_materia_ausente_en_un_trimestre_se_conserva_sin_desplazar_notas(self):
        t1 = self.crear_excel("materia_ausente_t1.xlsx", "3RO BGU", [("BIOLOGÍA", 9)])
        t2 = self.crear_excel("materia_ausente_t2.xlsx", "3RO BGU", [("MATEMÁTICA", 8.5), ("BIOLOGÍA", 9.25)])
        t3 = self.crear_excel("materia_ausente_t3.xlsx", "3RO BGU", [("MATEMÁTICA", 8.8), ("BIOLOGÍA", 9.4)])
        estudiante = consolidar_estudiantes(t1, t2, t3, None, "3RO BGU")[0]
        self.assertEqual(list(estudiante["materias"]), ["MATEMÁTICA", "BIOLOGÍA"])
        self.assertIsNone(estudiante["materias"]["MATEMÁTICA"]["t1"])
        self.assertEqual(estudiante["materias"]["BIOLOGÍA"]["t1"], 9.0)
        self.assertEqual(estudiante["materias"]["BIOLOGÍA"]["t2"], 9.25)

    def test_inyecta_comportamientos_distintos_y_respeta_vacio_en_html(self):
        plantilla = Path(__file__).parents[1] / "assets" / "certificados" / "FORMATO DE 3 DE BGU.html"
        html = inject_evaluacion_comportamental(
            plantilla.read_text(encoding="utf-8"),
            {"T1": "Valor T1", "T2": "", "T3": "Valor T3"},
        )
        self.assertIn("Valor T1", html)
        self.assertIn("Valor T3", html)
        self.assertNotIn("Transforma los desacuerdos en oportunidades de crecimiento y cooperación", html)
        self.assertNotRegex(html, r">\s*(None|nan|NaN|null|undefined)\s*<")

    def test_cursos_sin_formato_comportamental_permanecen_excluidos(self):
        for grado in ("INICIAL 1", "INICIAL 2", "1RO DE EGB"):
            with self.subTest(grado=grado):
                self.assertFalse(curso_admite_evaluacion_comportamental(grado))
        self.assertTrue(curso_admite_evaluacion_comportamental("3RO BGU"))

    def test_encabezado_desconocido_no_se_reconoce(self):
        self.assertEqual(clasificar_asignatura("TALLER DE ROBÓTICA", "1RO BGU")["estado"], "no_reconocida")

    def test_grado_distinto_rechaza_sin_alterar_datos_previos(self):
        ruta = self.crear_excel("tercero.xlsx", "3RO BGU", [("MATEMÁTICA SUPERIOR", 9)])
        datos_previos = {"curso": {"promedio": 8.5}}
        copia = {"curso": dict(datos_previos["curso"])}
        with self.assertRaisesRegex(ErrorGradoExcel, "El archivo corresponde a 3RO BGU"):
            cargar_excel_datos(ruta, "1RO BGU")
        self.assertEqual(datos_previos, copia)

    def test_materia_ignorada_no_participa_en_promedio(self):
        rutas = [
            self.crear_excel(f"p{n}.xlsx", "1RO BGU", [("MATEMÁTICA", 9), ("MATEMÁTICA SUPERIOR", 1)])
            for n in range(1, 4)
        ]
        resultado = consolidar_estudiantes(*rutas, None, "1RO BGU")[0]
        self.assertEqual(list(resultado["materias"]), ["MATEMÁTICA"])
        self.assertEqual(resultado["promedio"], 9)
        self.assertEqual(resultado["estado"], "APROBADO")

    def test_cedula_conserva_ceros_iniciales_segun_formato(self):
        ruta = self.crear_excel("cedula.xlsx", "1RO BGU", [("MATEMÁTICA", 9)], cedula=123456789)
        from openpyxl import load_workbook
        libro = load_workbook(ruta)
        libro["Reporte Periodo"]["B10"].number_format = "0000000000"
        libro.save(ruta)
        datos = cargar_excel_datos(ruta, "1RO BGU")
        self.assertIn("0123456789", datos)

    def test_nombres_exactos_existentes_siguen_funcionando(self):
        materias = ["MATEMÁTICA", "FÍSICA", "BIOLOGÍA", "QUÍMICA", "HISTORIA", "LENGUA Y LITERATURA", "INGLÉS", "EDUCACIÓN FÍSICA", "EMPRENDIMIENTO Y GESTIÓN"]
        for materia in materias:
            with self.subTest(materia=materia):
                self.assertEqual(clasificar_asignatura(materia, "3RO BGU")["estado"], "reconocida")

    def test_materias_especiales_de_tercero_no_se_extienden_a_primero(self):
        especiales = [
            "MATEMÁTICA SUPERIOR", "REDACCIÓN CREATIVA", "QUÍMICA SUPERIOR",
            "INVESTIGACIÓN EN CIENCIA Y TECNOLOGÍA", "SOCIOLOGÍA",
        ]
        for materia in especiales:
            with self.subTest(materia=materia):
                self.assertEqual(clasificar_asignatura(materia, "3RO BGU")["estado"], "reconocida")
                self.assertEqual(clasificar_asignatura(materia, "1RO BGU")["estado"], "ignorada_por_curso")

    def test_diagnostico_separa_reconocidas_ignoradas_y_desconocidas(self):
        ruta = self.crear_excel("diagnostico.xlsx", "1RO BGU", [("Bologia", 8), ("MATEMÁTICA SUPERIOR", 9), ("TALLER X", 7)])
        diagnostico = crear_diagnostico_asignaturas()
        datos = cargar_excel_datos(ruta, "1RO BGU", diagnostico)
        self.assertEqual(list(next(iter(datos.values()))["notas"]), ["BIOLOGÍA"])
        self.assertEqual(diagnostico["asignaturasReconocidas"][0]["canonica"], "BIOLOGÍA")
        self.assertEqual(diagnostico["asignaturasIgnoradasPorCurso"][0]["canonica"], "MATEMÁTICA SUPERIOR")
        self.assertEqual(diagnostico["asignaturasNoReconocidas"][0]["original"], "TALLER X")

    def test_equivalencias_de_grado(self):
        self.assertTrue(grados_equivalentes("1RO DE BACHILLERATO CIENCIAS", "PRIMER BGU"))
        self.assertFalse(grados_equivalentes("1RO BGU", "3RO BGU"))

    def test_cli_conserva_contrato_y_entrega_diagnostico(self):
        ruta = self.crear_excel("cli.xlsx", "PRIMER BGU", [("Bologia", 8), ("MATEMÁTICA SUPERIOR", 9)])
        proceso = subprocess.run(
            [sys.executable, str(Path(__file__).parents[1] / "procesador_notas.py"), "--analizar", "--grado", "1RO DE BACHILLERATO CIENCIAS", "--t1", str(ruta)],
            capture_output=True,
            text=True,
            encoding="utf-8",
            check=True,
        )
        respuesta = json.loads(proceso.stdout)
        self.assertIn("datosInstitucion", respuesta)
        self.assertIn("estudiantes", respuesta)
        self.assertEqual(respuesta["asignaturas"], ["BIOLOGÍA"])
        self.assertEqual(respuesta["asignaturasIgnoradasPorCurso"][0]["canonica"], "MATEMÁTICA SUPERIOR")

    def test_cli_rechaza_grado_distinto_con_mensaje_comprensible(self):
        ruta = self.crear_excel("cli_otro_grado.xlsx", "3RO BGU", [("MATEMÁTICA SUPERIOR", 9)])
        proceso = subprocess.run(
            [sys.executable, str(Path(__file__).parents[1] / "procesador_notas.py"), "--analizar", "--grado", "1RO BGU", "--t1", str(ruta)],
            capture_output=True,
            text=True,
            encoding="utf-8",
            check=True,
        )
        respuesta = json.loads(proceso.stdout)
        self.assertEqual(respuesta["error"], "El archivo corresponde a 3RO BGU, pero el curso seleccionado es 1RO BGU.")

    def test_catalogo_declara_las_16_optativas_con_metadatos_y_solo_para_bgu3(self):
        self.assertEqual(
            [entrada["nombre"] for entrada in CATALOGO_ASIGNATURAS if entrada.get("es_optativa_bgu3")],
            OPTATIVAS_BGU3,
        )
        for orden, materia in enumerate(OPTATIVAS_BGU3, start=12):
            with self.subTest(materia=materia):
                clasificacion = clasificar_asignatura(materia, "3RO BGU")
                self.assertEqual(clasificacion["estado"], "reconocida")
                self.assertEqual(clasificacion["tipo"], "cuantitativa")
                self.assertEqual(clasificacion["categoria"], "optativa")
                self.assertTrue(clasificacion["es_optativa_bgu3"])
                self.assertEqual(clasificacion["presentacion_certificado"], "escala_cualitativa")
                self.assertFalse(clasificacion["permite_supletorio"])
                self.assertEqual(clasificacion["orden"], orden * 10)
                self.assertEqual(clasificar_asignatura(materia, "1RO BGU")["estado"], "ignorada_por_curso")

        matematica = metadatos_asignatura("MATEMÁTICA")
        self.assertFalse(matematica["es_optativa_bgu3"])
        self.assertTrue(matematica["permite_supletorio"])
        self.assertNotEqual(matematica["presentacion_certificado"], "escala_cualitativa")
        superior = next(e for e in CATALOGO_ASIGNATURAS if e["nombre"] == "MATEMÁTICA SUPERIOR")
        self.assertNotIn("MATEMÁTICA", superior["aliases"])
        self.assertNotIn("MATEMATICA", superior["aliases"])

    def test_aliases_oficiales_se_resuelven_a_sus_nombres_canonicos(self):
        casos = {
            "APRECIACION MUSICAL": "APRECIACIÓN MUSICAL",
            "CORRIENTES FILOSOFICAS": "CORRIENTES FILOSÓFICAS",
            "INVESTIGACION EN CIENCIA Y TECNOLOGIA": "INVESTIGACIÓN EN CIENCIA Y TECNOLOGÍA",
            "LECTURA CRITICA DE MENSAJES": "LECTURA CRÍTICA DE MENSAJES",
            "KICHWA": "LENGUA KICHWA",
            "REDACCION CREATIVA": "REDACCIÓN CREATIVA",
            "SOCIOLOGIA": "SOCIOLOGÍA",
            "ARTE TEATRAL": "TEATRO",
            "PROBLEMAS DEL MUNDO CONTEMPORANEO": "PROBLEMAS DEL MUNDO CONTEMPORÁNEO",
            "PSICOLOGIA": "PSICOLOGÍA",
            "LENGUA EXTRANJERA FRANCES": "LENGUA EXTRANJERA: FRANCÉS",
            "FISICA SUPERIOR": "FÍSICA SUPERIOR",
            "QUIMICA SUPERIOR": "QUÍMICA SUPERIOR",
            "MATEMATICA SUPERIOR": "MATEMÁTICA SUPERIOR",
            "MATEMATICAS SUPERIORES": "MATEMÁTICA SUPERIOR",
            "BIOLOGIA SUPERIOR": "BIOLOGÍA SUPERIOR",
            "LENGUA Y CULTURA ANCESTRAL": "LENGUA Y CULTURA ANCESTRAL",
        }
        for alias, canonica in casos.items():
            with self.subTest(alias=alias):
                resultado = clasificar_asignatura(alias, "3RO BGU")
                self.assertEqual(resultado["estado"], "reconocida")
                self.assertEqual(resultado["canonica"], canonica)

    def test_nombres_obligatorios_no_se_confunden_con_optativas_parecidas(self):
        casos = {
            "FÍSICA": "FÍSICA",
            "QUÍMICA": "QUÍMICA",
            "BIOLOGÍA": "BIOLOGÍA",
            "MATEMÁTICA": "MATEMÁTICA",
            "MATEMATICA": "MATEMÁTICA",
            "MATEMÁTICA SUPERIOR": "MATEMÁTICA SUPERIOR",
            "LENGUA Y LITERATURA": "LENGUA Y LITERATURA",
            "INGLÉS": "INGLÉS",
            "LECTURA CRÍTICA": "LECTURA CRÍTICA DE MENSAJES",
        }
        for original, canonica in casos.items():
            with self.subTest(original=original):
                self.assertEqual(clasificar_asignatura(original, "3RO BGU")["canonica"], canonica)

        filosofia = clasificar_asignatura("FILOSOFÍA", "3RO BGU")
        self.assertEqual(filosofia["canonica"], "FILOSOFÍA")
        self.assertEqual(filosofia["estado"], "ignorada_por_curso")
        self.assertNotEqual(filosofia["canonica"], "CORRIENTES FILOSÓFICAS")

    def test_escala_cualitativa_respeta_todos_los_limites_y_vacios(self):
        casos = [
            (1.00, "E-"), (1.49, "E-"), (1.50, "E+"), (2.49, "E+"),
            (2.50, "D-"), (3.49, "D-"), (3.50, "D+"), (4.49, "D+"),
            (4.50, "C-"), (5.49, "C-"), (5.50, "C+"), (6.49, "C+"),
            (6.50, "B-"), (7.49, "B-"), (7.50, "B+"), (8.49, "B+"),
            (8.50, "A-"), (9.49, "A-"), (9.50, "A+"), (10.00, "A+"),
            (None, ""), ("", ""), (float("nan"), ""), (0.99, ""), (10.01, ""),
        ]
        for nota, esperado in casos:
            with self.subTest(nota=nota):
                self.assertEqual(convertir_nota_optativa_a_escala_cualitativa(nota), esperado)

    def test_metadatos_viajan_desde_encabezado_hasta_materia_consolidada(self):
        rutas = [
            self.crear_excel(f"meta_t{periodo}.xlsx", "3RO BGU", [("APRECIACION MUSICAL", 9 + periodo / 10)])
            for periodo in range(1, 4)
        ]
        registro = next(iter(cargar_excel_datos(rutas[0], "3RO BGU").values()))
        metadata_excel = registro["metadatos_asignaturas"]["APRECIACIÓN MUSICAL"]
        self.assertTrue(metadata_excel["es_optativa_bgu3"])
        self.assertFalse(metadata_excel["permite_supletorio"])

        materia = consolidar_estudiantes(*rutas, None, "3RO BGU")[0]["materias"]["APRECIACIÓN MUSICAL"]
        for campo in ("tipo", "categoria", "es_optativa_bgu3", "presentacion_certificado", "permite_supletorio", "orden"):
            self.assertEqual(materia[campo], metadata_excel[campo])
        self.assertEqual((materia["t1"], materia["t2"], materia["t3"]), (9.1, 9.2, 9.3))
        self.assertIsNone(materia["estado"])

    def test_matematica_y_matematica_superior_coexisten_sin_mezclar_notas_ni_estado(self):
        rutas = [
            self.crear_excel(
                f"coexistencia_t{periodo}.xlsx",
                "3RO BGU",
                [("MATEMÁTICA", 6.0), ("MATEMÁTICA SUPERIOR", 9.55)],
            )
            for periodo in range(1, 4)
        ]
        estudiante = consolidar_estudiantes(*rutas, None, "3RO BGU")[0]
        self.assertEqual(list(estudiante["materias"])[:2], ["MATEMÁTICA", "MATEMÁTICA SUPERIOR"])
        matematica = estudiante["materias"]["MATEMÁTICA"]
        superior = estudiante["materias"]["MATEMÁTICA SUPERIOR"]
        self.assertEqual((matematica["t1"], matematica["t2"], matematica["t3"]), (6.0, 6.0, 6.0))
        self.assertEqual(matematica["estado"], "SUPLETORIO")
        self.assertTrue(matematica["permite_supletorio"])
        self.assertEqual((superior["t1"], superior["t2"], superior["t3"]), (9.55, 9.55, 9.55))
        self.assertIsNone(superior["estado"])
        self.assertFalse(superior["permite_supletorio"])
        self.assertEqual(estudiante["estado"], "SUPLETORIO")
        self.assertEqual(estudiante["promedio"], 6.0)

        plantilla = (Path(__file__).parents[1] / "assets" / "certificados" / "FORMATO DE 3 DE BGU.html").read_text(encoding="utf-8")
        html = inject_subject_grades(plantilla, estudiante["materias"])
        optativas = self.contenido_optativas(html)
        self.assertIn("MATEMÁTICA SUPERIOR", optativas)
        self.assertIn(">A+</td>", optativas)
        self.assertNotRegex(optativas, r">MATEMÁTICA</td>")

    def test_union_de_optativas_t1_t2_t3_no_copia_ni_desplaza_valores(self):
        t1 = self.crear_excel("union_t1.xlsx", "3RO BGU", [("TEATRO", 8.1), ("PSICOLOGÍA", 6.8)])
        t2 = self.crear_excel("union_t2.xlsx", "3RO BGU", [("TEATRO", 8.2), ("PSICOLOGÍA", 6.9), ("APRECIACIÓN MUSICAL", 9.7)])
        t3 = self.crear_excel("union_t3.xlsx", "3RO BGU", [("TEATRO", 8.3), ("APRECIACIÓN MUSICAL", 9.6)])
        estudiante = consolidar_estudiantes(t1, t2, t3, None, "3RO BGU")[0]
        self.assertEqual(
            [nombre for nombre, datos in estudiante["materias"].items() if datos["es_optativa_bgu3"]],
            ["APRECIACIÓN MUSICAL", "TEATRO", "PSICOLOGÍA"],
        )
        self.assertEqual(
            (estudiante["materias"]["PSICOLOGÍA"]["t1"], estudiante["materias"]["PSICOLOGÍA"]["t2"], estudiante["materias"]["PSICOLOGÍA"]["t3"]),
            (6.8, 6.9, None),
        )
        self.assertEqual(
            (estudiante["materias"]["APRECIACIÓN MUSICAL"]["t1"], estudiante["materias"]["APRECIACIÓN MUSICAL"]["t2"], estudiante["materias"]["APRECIACIÓN MUSICAL"]["t3"]),
            (None, 9.7, 9.6),
        )
        plantilla = (Path(__file__).parents[1] / "assets" / "certificados" / "FORMATO DE 3 DE BGU.html").read_text(encoding="utf-8")
        optativas = self.contenido_optativas(inject_subject_grades(plantilla, estudiante["materias"]))
        self.assertRegex(optativas, r"PSICOLOGÍA[\s\S]*?<td>B-</td>[\s\S]*?<td>B-</td>[\s\S]*?<td></td>")
        self.assertRegex(optativas, r"APRECIACIÓN MUSICAL[\s\S]*?<td></td>[\s\S]*?<td>A\+</td>[\s\S]*?<td>A\+</td>")

    def test_tres_combinaciones_institucionales_generan_solo_sus_optativas(self):
        instituciones = [
            {
                "MATEMÁTICA SUPERIOR": (9.55, "A+"),
                "REDACCIÓN CREATIVA": (7.76, "B+"),
                "QUÍMICA SUPERIOR": (8.40, "B+"),
                "INVESTIGACIÓN EN CIENCIA Y TECNOLOGÍA": (8.18, "B+"),
                "SOCIOLOGÍA": (9.41, "A-"),
            },
            {
                "APRECIACIÓN MUSICAL": (9.60, "A+"),
                "TEATRO": (8.30, "B+"),
                "PSICOLOGÍA": (6.70, "B-"),
                "LECTURA CRÍTICA DE MENSAJES": (5.80, "C+"),
            },
            {
                "LENGUA KICHWA": (7.50, "B+"),
                "LENGUA Y CULTURA ANCESTRAL": (8.50, "A-"),
                "PROBLEMAS DEL MUNDO CONTEMPORÁNEO": (4.50, "C-"),
                "CORRIENTES FILOSÓFICAS": (2.50, "D-"),
            },
        ]
        plantilla = (Path(__file__).parents[1] / "assets" / "certificados" / "FORMATO DE 3 DE BGU.html").read_text(encoding="utf-8")
        for indice, institucion in enumerate(instituciones, start=1):
            with self.subTest(institucion=indice):
                materias = {
                    nombre: {"tipo": "cuantitativa", "t1": nota, "t2": None, "t3": None}
                    for nombre, (nota, _) in institucion.items()
                }
                optativas = self.contenido_optativas(inject_optativas_bgu3(plantilla, materias))
                for nombre, (_, escala) in institucion.items():
                    self.assertIn(nombre, optativas)
                    self.assertIn(f"<td>{escala}</td>", optativas)
                for ausente in set(OPTATIVAS_BGU3) - set(institucion):
                    self.assertNotIn(ausente, optativas)

    def test_optativas_bajas_no_generan_supletorio_reprobacion_ni_registro(self):
        materias_bajas = [(nombre, 2.0 + (indice % 3)) for indice, nombre in enumerate(OPTATIVAS_BGU3)]
        rutas = [
            self.crear_excel(f"bajas_t{periodo}.xlsx", "3RO BGU", [("MATEMÁTICA", 6.0), *materias_bajas])
            for periodo in range(1, 4)
        ]
        estudiante = consolidar_estudiantes(*rutas, None, "3RO BGU")[0]
        for nombre in OPTATIVAS_BGU3:
            with self.subTest(materia=nombre):
                materia = estudiante["materias"][nombre]
                self.assertIsNone(materia["estado"])
                self.assertIsNone(materia["supletorio"])
                self.assertFalse(materia["permite_supletorio"])
                self.assertFalse(actualizar_certificado_supletorio("sin_archivo", nombre, 10, self.directorio))
        self.assertEqual(estudiante["materias"]["MATEMÁTICA"]["estado"], "SUPLETORIO")
        self.assertEqual(estudiante["estado"], "SUPLETORIO")

        resultado = generar_certificados_inicial({
            "datos_consolidados": [estudiante],
            "institucion": {"grado": "3RO BGU", "nivel": "BGU"},
            "logos": {},
            "gradoCursoCanonico": "3RO BGU",
            "plantillaName": "FORMATO DE 3 DE BGU.html",
            "certOutputDir": str(self.directorio / "certificados_optativas_bajas"),
        })
        self.assertEqual([item["asignatura"] for item in resultado["supletorios"]], ["MATEMÁTICA"])
        html = Path(resultado["archivos"][0]).read_text(encoding="utf-8")
        optativas = self.contenido_optativas(html)
        self.assertEqual(optativas.count("<tr>"), 16)
        self.assertNotIn(">2.00</td>", optativas)

    def test_una_futura_optativa_funciona_agregando_solo_una_entrada_al_catalogo(self):
        nueva = {
            "nombre": "NUEVA OPTATIVA OFICIAL",
            "grados": ["BGU_3"],
            "aliases": ["NUEVA OPTATIVA"],
            "tipo": "cuantitativa",
            "categoria": "optativa",
            "es_optativa_bgu3": True,
            "presentacion_certificado": "escala_cualitativa",
            "permite_supletorio": False,
            "orden": 400,
        }
        CATALOGO_ASIGNATURAS.append(nueva)
        try:
            clasificacion = clasificar_asignatura("NUEVA OPTATIVA", "3RO BGU")
            self.assertEqual(clasificacion["canonica"], "NUEVA OPTATIVA OFICIAL")
            plantilla = (Path(__file__).parents[1] / "assets" / "certificados" / "FORMATO DE 3 DE BGU.html").read_text(encoding="utf-8")
            optativas = self.contenido_optativas(inject_optativas_bgu3(
                plantilla,
                {"NUEVA OPTATIVA OFICIAL": {"t1": 9.8, "t2": 8.8, "t3": 7.8}},
            ))
            self.assertIn("NUEVA OPTATIVA OFICIAL", optativas)
            self.assertRegex(optativas, r"<td>A\+</td>[\s\S]*<td>A-</td>[\s\S]*<td>B\+</td>")
        finally:
            CATALOGO_ASIGNATURAS.remove(nueva)

    def test_diagnostico_separa_conteo_de_optativas_sin_duplicarlas(self):
        ruta = self.crear_excel(
            "diagnostico_optativas.xlsx",
            "3RO BGU",
            [("MATEMÁTICA", 8), ("APRECIACION MUSICAL", 9), ("TEATRO", 8)],
        )
        diagnostico = crear_diagnostico_asignaturas()
        cargar_excel_datos(ruta, "3RO BGU", diagnostico)
        self.assertEqual(
            [item["canonica"] for item in diagnostico["optativasDetectadas"]],
            ["APRECIACIÓN MUSICAL", "TEATRO"],
        )


if __name__ == "__main__":
    unittest.main()
