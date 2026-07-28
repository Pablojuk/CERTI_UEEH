# -*- coding: utf-8 -*-

import hashlib
import json
import shutil
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from catalogo_asignaturas import CATALOGO_ASIGNATURAS, normalizar_texto_asignatura
from procesador_notas import (
    cargar_excel_datos,
    consolidar_estudiantes,
    generar_formato_bgu3_con_optativas,
)


class FormatoBgu3OptativasTests(unittest.TestCase):
    def setUp(self):
        self.temporal = tempfile.TemporaryDirectory()
        self.directorio = Path(self.temporal.name)
        self.origen = (
            Path(__file__).parents[1]
            / "assets"
            / "formatos-notas"
            / "formato_3_bgu.xlsx"
        )
        self.optativas = [
            "APRECIACIÓN MUSICAL",
            "INVESTIGACIÓN EN CIENCIA Y TECNOLOGÍA",
            "TEATRO",
        ]
        self.optativas_catalogo = {
            entrada["nombre"]
            for entrada in CATALOGO_ASIGNATURAS
            if entrada.get("es_optativa_bgu3") is True
        }

    def tearDown(self):
        self.temporal.cleanup()

    @staticmethod
    def hash_archivo(ruta):
        return hashlib.sha256(Path(ruta).read_bytes()).hexdigest()

    @staticmethod
    def encabezados_materias(hoja):
        return [
            str(hoja.cell(9, columna).value).strip()
            for columna in range(3, 19)
            if hoja.cell(9, columna).value not in (None, "")
        ]

    @staticmethod
    def validaciones(hoja):
        return sorted(
            (
                validacion.type,
                validacion.formula1,
                str(validacion.sqref),
            )
            for validacion in hoja.data_validations.dataValidation
        )

    def generar(self, nombre="generado.xlsx"):
        destino = self.directorio / nombre
        resultado = generar_formato_bgu3_con_optativas(
            self.origen,
            destino,
            self.optativas,
        )
        self.assertTrue(resultado["success"])
        return destino

    def test_tres_optativas_aparecen_exactamente_y_la_plantilla_original_no_cambia(self):
        hash_original = self.hash_archivo(self.origen)
        libro_original = load_workbook(self.origen)
        hoja_original = libro_original.active
        encabezados_originales = self.encabezados_materias(hoja_original)
        materias_fijas = [
            nombre
            for nombre in encabezados_originales
            if nombre not in self.optativas_catalogo
        ]
        validaciones_originales = self.validaciones(hoja_original)

        destino = self.generar()
        libro_generado = load_workbook(destino)
        hoja_generada = libro_generado.active
        encabezados = self.encabezados_materias(hoja_generada)

        self.assertEqual(
            [nombre for nombre in encabezados if nombre in self.optativas_catalogo],
            self.optativas,
        )
        self.assertEqual(
            [nombre for nombre in encabezados if nombre not in self.optativas_catalogo],
            materias_fijas,
        )
        self.assertEqual(len(encabezados), len(materias_fijas) + 3)
        self.assertLessEqual(len(encabezados), 16)
        self.assertTrue(all(
            hoja_generada.cell(9, columna).value in (None, "")
            for columna in range(3 + len(encabezados), 19)
        ))
        self.assertEqual(self.validaciones(hoja_generada), validaciones_originales)
        self.assertEqual(self.hash_archivo(self.origen), hash_original)

        estilos_originales = {
            normalizar_texto_asignatura(hoja_original.cell(9, columna).value): (
                hoja_original.cell(9, columna).style_id,
                hoja_original.cell(10, columna).style_id,
                hoja_original.column_dimensions[get_column_letter(columna)].width,
            )
            for columna in range(3, 19)
            if hoja_original.cell(9, columna).value
        }
        for columna in range(3, 3 + len(encabezados)):
            nombre = hoja_generada.cell(9, columna).value
            normalizado = normalizar_texto_asignatura(nombre)
            if nombre in self.optativas:
                continue
            self.assertEqual(
                (
                    hoja_generada.cell(9, columna).style_id,
                    hoja_generada.cell(10, columna).style_id,
                    hoja_generada.column_dimensions[get_column_letter(columna)].width,
                ),
                estilos_originales[normalizado],
            )

    def test_sin_optativas_y_sobre_capacidad_bloquean_la_generacion(self):
        with self.assertRaisesRegex(ValueError, "al menos una"):
            generar_formato_bgu3_con_optativas(
                self.origen,
                self.directorio / "sin_optativas.xlsx",
                [],
            )
        seis_optativas = [
            entrada["nombre"]
            for entrada in CATALOGO_ASIGNATURAS
            if entrada.get("es_optativa_bgu3") is True
        ][:6]
        with self.assertRaisesRegex(ValueError, "capacidad"):
            generar_formato_bgu3_con_optativas(
                self.origen,
                self.directorio / "exceso.xlsx",
                seis_optativas,
            )

    def test_el_formato_generado_se_procesa_en_t1_t2_y_t3(self):
        base = self.generar("base.xlsx")
        rutas = []
        for numero, periodo in enumerate(
            ("TRIMESTRE 1", "TRIMESTRE 2", "TRIMESTRE 3"),
            start=1,
        ):
            ruta = self.directorio / f"t{numero}.xlsx"
            shutil.copy2(base, ruta)
            libro = load_workbook(ruta)
            hoja = libro.active
            hoja["B6"] = periodo
            hoja["A10"] = "ESTUDIANTE PRUEBA"
            hoja["B10"] = "0123456789"
            for columna in range(3, 19):
                encabezado = hoja.cell(9, columna).value
                if not encabezado:
                    continue
                normalizado = normalizar_texto_asignatura(encabezado)
                if normalizado == "EVALUACION COMPORTAMENTAL":
                    hoja.cell(10, columna).value = f"Comportamiento T{numero}"
                elif "CIVICA Y ACOMPANAMIENTO" in normalizado:
                    hoja.cell(10, columna).value = "B+"
                else:
                    hoja.cell(10, columna).value = 8 + numero / 10
            libro.save(ruta)
            rutas.append(ruta)

            estudiante = cargar_excel_datos(ruta, "3RO BGU")["0123456789"]
            optativas_detectadas = [
                nombre
                for nombre in estudiante["notas"]
                if nombre in self.optativas_catalogo
            ]
            self.assertEqual(optativas_detectadas, self.optativas)

        consolidado = consolidar_estudiantes(
            rutas[0],
            rutas[1],
            rutas[2],
            None,
            "3RO BGU",
        )[0]
        for optativa in self.optativas:
            self.assertEqual(
                (
                    consolidado["materias"][optativa]["t1"],
                    consolidado["materias"][optativa]["t2"],
                    consolidado["materias"][optativa]["t3"],
                ),
                (8.1, 8.2, 8.3),
            )

    def test_cli_genera_la_copia_solicitada_por_electron(self):
        destino = self.directorio / "desde_cli.xlsx"
        proceso = subprocess.run(
            [
                sys.executable,
                str(Path(__file__).parents[1] / "procesador_notas.py"),
                "--generar-formato-bgu3",
            ],
            input=json.dumps({
                "origen": str(self.origen),
                "destino": str(destino),
                "optativas": self.optativas,
            }),
            capture_output=True,
            text=True,
            encoding="utf-8",
            check=True,
        )
        respuesta = json.loads(proceso.stdout)
        self.assertTrue(respuesta["success"])
        self.assertEqual(respuesta["optativas"], self.optativas)
        self.assertTrue(destino.is_file())


if __name__ == "__main__":
    unittest.main()
